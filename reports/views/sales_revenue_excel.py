"""
Savdo tushumlari (faktura + OKKM) — tashkilotlar bo'yicha OYLIK Excel hisobot.

Har bir tashkilot (ShopTenant) qator, har oy (Yan..Dek) ustun. Qiymat =
FacturaRevenueDaily.sales + OkkmRevenueDaily.turnover (xom so'm), shu oyda
yig'ilgan. Bu dashboarddagi "Savdo tushumlari" bilan bir xil manba.

Identifikator ustuni tashkilot turiga qarab:
  - YTT  -> PINFL (rahbar JSHSHIR)
  - MCHJ -> STIR
Bu sync (revenue_sync._resolve_tin) bilan bir mantiq: YTT pinfl bo'yicha,
MCHJ stir bo'yicha olinadi.

GET /api/v1/reports/sales-revenue/excel?year=2026
  ?year= berilmasa — joriy yil.
"""
import io
from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from monitoring.models import FacturaRevenueDaily, OkkmRevenueDaily, ShopTenant

MONTHS_UZ = [
    "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
    "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr",
]

TYPE_LABEL = {
    ShopTenant.BusinessType.YTT: "YTT",
    ShopTenant.BusinessType.LEGAL: "MCHJ",
    ShopTenant.BusinessType.OTHER: "Boshqa",
}

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


class SalesRevenueExcelView(APIView):
    """Tashkilotlar bo'yicha oylik savdo tushumlari (faktura + OKKM) — .xlsx."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            year = int(request.query_params.get("year") or date.today().year)
        except (TypeError, ValueError):
            year = date.today().year

        # ?source=factura | okkm | both (default both)
        source = (request.query_params.get("source") or "both").lower()
        if source not in ("factura", "okkm", "both"):
            source = "both"

        # 1) Tanlangan manba(lar)ni tin x oy bo'yicha DB da yig'amiz.
        #    revenue[tin][month_index 0..11] = Decimal savdo
        revenue: dict[str, list[Decimal]] = defaultdict(
            lambda: [Decimal(0)] * 12
        )

        if source in ("factura", "both"):
            for r in (
                FacturaRevenueDaily.objects
                .filter(date__year=year)
                .values("seller_tin", "date__month")
                .annotate(s=Sum("sales"))
            ):
                tin = (r["seller_tin"] or "").strip()
                if tin:
                    revenue[tin][r["date__month"] - 1] += r["s"] or Decimal(0)

        if source in ("okkm", "both"):
            for r in (
                OkkmRevenueDaily.objects
                .filter(date__year=year)
                .values("tin", "date__month")
                .annotate(o=Sum("turnover"))
            ):
                tin = (r["tin"] or "").strip()
                if tin:
                    revenue[tin][r["date__month"] - 1] += r["o"] or Decimal(0)

        # 2) tin (= revenue kaliti, odatda stir) -> tashkilot ma'lumoti. Bitta tin
        #    bir nechta do'konda bo'lishi mumkin — nom/do'konlar birlashtiriladi.
        info_by_tin: dict[str, dict] = {}
        for t in (
            ShopTenant.objects
            .exclude(stir__isnull=True).exclude(stir="")
            .select_related("shop")
        ):
            stir = (t.stir or "").strip()
            if not stir:
                continue
            entry = info_by_tin.setdefault(stir, {
                "name": None, "shops": [], "btype": None,
                "stir": stir, "pinfl": None,
            })
            if not entry["name"]:
                entry["name"] = (t.name or t.leader_fio or "").strip() or None
            if not entry["btype"]:
                entry["btype"] = t.business_type
            if not entry["pinfl"]:
                entry["pinfl"] = (t.leader_jshshir or "").strip() or None
            if t.shop_id:
                label = str(t.shop)
                if label not in entry["shops"]:
                    entry["shops"].append(label)

        # 3) Qatorlar: jami bo'yicha kamayish tartibida (eng katta tashkilot tepada).
        rows = []
        for tin, months in revenue.items():
            total = sum(months, Decimal(0))
            if total == 0:
                continue  # shu yilda savdosi bo'lmagan tin lar tushmasin
            info = info_by_tin.get(tin)
            name, type_label, ident, shop = self._identity(tin, info)
            rows.append({
                "name": name,
                "type": type_label,
                "ident": ident,
                "shop": shop,
                "months": months,
                "total": total,
            })
        rows.sort(key=lambda x: x["total"], reverse=True)

        wb = self._build_workbook(year, rows, source)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
        resp["Content-Disposition"] = (
            f'attachment; filename="savdo_tushumlari_{source}_{year}.xlsx"'
        )
        return resp

    # ------------------------------------------------------------------
    @staticmethod
    def _identity(tin: str, info: dict | None):
        """
        (nom, turi, identifikator, do'kon) ni qaytaradi.
          - YTT  -> identifikator = PINFL
          - MCHJ -> identifikator = STIR
        Tashkilot topilmasa — bari tin.
        """
        if not info:
            return tin, "", tin, ""
        btype = info["btype"]
        if btype == ShopTenant.BusinessType.YTT and info["pinfl"]:
            ident = info["pinfl"]
        else:
            ident = info["stir"] or tin
        type_label = TYPE_LABEL.get(btype, "")
        name = info["name"] or ident
        shop = ", ".join(info["shops"]) if info["shops"] else ""
        return name, type_label, ident, shop

    # ------------------------------------------------------------------
    def _build_workbook(self, year: int, rows: list[dict], source: str = "both") -> Workbook:
        source_label = {
            "factura": "faktura",
            "okkm": "OKKM",
            "both": "faktura + OKKM",
        }.get(source, "faktura + OKKM")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill("solid", fgColor="DDEBF7")
        money_fmt = "#,##0"

        # Sobit ustunlar: №, Tashkilot, Turi, STIR / PINFL, Do'kon
        fixed = ["№", "Tashkilot", "Turi", "STIR / PINFL", "Do'kon"]
        n_fixed = len(fixed)
        month_col0 = n_fixed + 1          # birinchi oy ustuni (1-based)
        total_col = month_col0 + 12       # "Jami" ustuni
        headers = fixed + MONTHS_UZ + ["Jami"]
        last_col = len(headers)

        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}"

        # Sarlavha (merged)
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=last_col
        )
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = (
            f"Savdo tushumlari ({source_label}) — {year}-yil, "
            f"tashkilotlar bo'yicha oylik (so'm)"
        )
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Ustun sarlavhalari
        for col, text in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=text)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            c.border = border
        ws.row_dimensions[2].height = 30

        col_totals = [Decimal(0)] * 12
        grand_total = Decimal(0)

        row_idx = 3
        for i, r in enumerate(rows, start=1):
            ws.cell(row=row_idx, column=1, value=i).border = border
            ws.cell(row=row_idx, column=2, value=r["name"]).border = border
            ws.cell(row=row_idx, column=3, value=r["type"]).border = border
            # Identifikator (STIR/PINFL) — matn sifatida (uzun raqam buzilmasin)
            ic = ws.cell(row=row_idx, column=4, value=r["ident"])
            ic.number_format = "@"
            ic.border = border
            ws.cell(row=row_idx, column=5, value=r["shop"]).border = border
            for m in range(12):
                val = r["months"][m]
                c = ws.cell(row=row_idx, column=month_col0 + m, value=float(val))
                c.number_format = money_fmt
                c.border = border
                col_totals[m] += val
            tc = ws.cell(row=row_idx, column=total_col, value=float(r["total"]))
            tc.number_format = money_fmt
            tc.font = Font(bold=True)
            tc.border = border
            grand_total += r["total"]
            row_idx += 1

        # Jami qatori
        for col in range(1, n_fixed + 1):
            c = ws.cell(row=row_idx, column=col, value="JAMI" if col == 2 else "")
            c.fill = total_fill
            c.font = Font(bold=True)
            c.border = border
        for m in range(12):
            c = ws.cell(row=row_idx, column=month_col0 + m, value=float(col_totals[m]))
            c.number_format = money_fmt
            c.font = Font(bold=True)
            c.fill = total_fill
            c.border = border
        gc = ws.cell(row=row_idx, column=total_col, value=float(grand_total))
        gc.number_format = money_fmt
        gc.font = Font(bold=True)
        gc.fill = total_fill
        gc.border = border

        # Ustun kengliklari
        widths = {1: 5, 2: 32, 3: 8, 4: 16, 5: 22}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        for m in range(12):
            ws.column_dimensions[get_column_letter(month_col0 + m)].width = 13
        ws.column_dimensions[get_column_letter(total_col)].width = 16

        # Sarlavhalarni va sobit ustunlarni muzlatish (birinchi oy ustunidan boshlab oqar)
        ws.freeze_panes = f"{get_column_letter(month_col0)}3"

        return wb
