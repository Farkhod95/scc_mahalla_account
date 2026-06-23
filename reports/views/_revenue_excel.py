"""
Tashkilotlar bo'yicha OYLIK tushum Excel hisobotlari uchun umumiy yordamchilar.

Ikkala hisobot bir xil ko'rinishda: har bir tashkilot — qator, Yan..Dek — ustun,
oxirida "Jami". Savdo tushumi (sales_revenue_excel) va Soliq tushumi
(tax_revenue_excel) shu yerdagi build_monthly_workbook / collect_tenant_info /
rows_from_revenue / xlsx_response ni ulashadi.

Identifikator ustuni tashkilot turiga qarab:
  - YTT  -> PINFL (rahbar JSHSHIR)
  - MCHJ -> STIR
"""
import io
from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from monitoring.models import ShopTenant

MONTHS_UZ = [
    "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
    "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr",
]

TYPE_LABEL = {
    ShopTenant.BusinessType.YTT: "YTT",
    ShopTenant.BusinessType.LEGAL: "MCHJ",
    ShopTenant.BusinessType.OTHER: "Boshqa",
}

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def collect_tenant_info() -> dict:
    """
    stir -> {name, shops, btype, stir, pinfl}. Bitta tin (stir) bir nechta do'konda
    bo'lishi mumkin — nom/do'konlar birlashtiriladi.
    """
    info_by_tin: dict[str, dict] = {}
    for t in (
        ShopTenant.objects
        .exclude(stir__isnull=True).exclude(stir="")
        .select_related("shop")
    ):
        stir = (t.stir or "").strip()
        if not stir:
            continue
        entry = info_by_tin.setdefault(stir, {
            "name": None, "shops": [], "btype": None,
            "stir": stir, "pinfl": None,
        })
        if not entry["name"]:
            entry["name"] = (t.name or t.leader_fio or "").strip() or None
        if not entry["btype"]:
            entry["btype"] = t.business_type
        if not entry["pinfl"]:
            entry["pinfl"] = (t.leader_jshshir or "").strip() or None
        if t.shop_id:
            label = str(t.shop)
            if label not in entry["shops"]:
                entry["shops"].append(label)
    return info_by_tin


def _identity(tin: str, info: dict | None):
    """
    (nom, turi, identifikator, do'kon) ni qaytaradi.
      - YTT  -> identifikator = PINFL
      - MCHJ -> identifikator = STIR
    Tashkilot topilmasa — bari tin.
    """
    if not info:
        return tin, "", tin, ""
    btype = info["btype"]
    if btype == ShopTenant.BusinessType.YTT and info["pinfl"]:
        ident = info["pinfl"]
    else:
        ident = info["stir"] or tin
    type_label = TYPE_LABEL.get(btype, "")
    name = info["name"] or ident
    shop = ", ".join(info["shops"]) if info["shops"] else ""
    return name, type_label, ident, shop


def rows_from_revenue(revenue: dict, info_by_tin: dict) -> list[dict]:
    """
    revenue[tin] = [12 ta Decimal] (oylik) -> jami bo'yicha kamayish tartibida
    tartiblangan qatorlar. Shu yilda tushumi 0 bo'lgan tin lar tushmaydi.
    """
    rows = []
    for tin, months in revenue.items():
        total = sum(months, Decimal(0))
        if total == 0:
            continue
        name, type_label, ident, shop = _identity(tin, info_by_tin.get(tin))
        rows.append({
            "name": name,
            "type": type_label,
            "ident": ident,
            "shop": shop,
            "months": months,
            "total": total,
        })
    rows.sort(key=lambda x: x["total"], reverse=True)
    return rows


def build_monthly_workbook(year: int, rows: list[dict], title: str) -> Workbook:
    """Tashkilot × oy jadvalini .xlsx Workbook qilib quradi (sarlavha — `title`)."""
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="DDEBF7")
    money_fmt = "#,##0"

    # Sobit ustunlar: №, Tashkilot, Turi, STIR / PINFL, Do'kon
    fixed = ["№", "Tashkilot", "Turi", "STIR / PINFL", "Do'kon"]
    n_fixed = len(fixed)
    month_col0 = n_fixed + 1          # birinchi oy ustuni (1-based)
    total_col = month_col0 + 12       # "Jami" ustuni
    headers = fixed + MONTHS_UZ + ["Jami"]
    last_col = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}"

    # Sarlavha (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Ustun sarlavhalari
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=text)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 30

    col_totals = [Decimal(0)] * 12
    grand_total = Decimal(0)

    row_idx = 3
    for i, r in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=i).border = border
        ws.cell(row=row_idx, column=2, value=r["name"]).border = border
        ws.cell(row=row_idx, column=3, value=r["type"]).border = border
        # Identifikator (STIR/PINFL) — matn sifatida (uzun raqam buzilmasin)
        ic = ws.cell(row=row_idx, column=4, value=r["ident"])
        ic.number_format = "@"
        ic.border = border
        ws.cell(row=row_idx, column=5, value=r["shop"]).border = border
        for m in range(12):
            val = r["months"][m]
            c = ws.cell(row=row_idx, column=month_col0 + m, value=float(val))
            c.number_format = money_fmt
            c.border = border
            col_totals[m] += val
        tc = ws.cell(row=row_idx, column=total_col, value=float(r["total"]))
        tc.number_format = money_fmt
        tc.font = Font(bold=True)
        tc.border = border
        grand_total += r["total"]
        row_idx += 1

    # Jami qatori
    for col in range(1, n_fixed + 1):
        c = ws.cell(row=row_idx, column=col, value="JAMI" if col == 2 else "")
        c.fill = total_fill
        c.font = Font(bold=True)
        c.border = border
    for m in range(12):
        c = ws.cell(row=row_idx, column=month_col0 + m, value=float(col_totals[m]))
        c.number_format = money_fmt
        c.font = Font(bold=True)
        c.fill = total_fill
        c.border = border
    gc = ws.cell(row=row_idx, column=total_col, value=float(grand_total))
    gc.number_format = money_fmt
    gc.font = Font(bold=True)
    gc.fill = total_fill
    gc.border = border

    # Ustun kengliklari
    widths = {1: 5, 2: 32, 3: 8, 4: 16, 5: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for m in range(12):
        ws.column_dimensions[get_column_letter(month_col0 + m)].width = 13
    ws.column_dimensions[get_column_letter(total_col)].width = 16

    # Sarlavhalarni va sobit ustunlarni muzlatish (birinchi oy ustunidan oqar)
    ws.freeze_panes = f"{get_column_letter(month_col0)}3"

    return wb


def xlsx_response(wb: Workbook, filename: str) -> HttpResponse:
    """Workbook ni .xlsx attachment HttpResponse qilib qaytaradi."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
