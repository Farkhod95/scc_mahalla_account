import os
from collections import OrderedDict

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from monitoring.models import Shop, ShopTenant, TenantEmployee
from monitoring.tasks import (
    empty_or_zero_to_none,
    normalize_block_type,
    normalize_business_type,
    normalize_fire_safety_level,
    normalize_activity_status,
    normalize_jshshir,
    normalize_phone,
    normalize_shop_number,
    normalize_tax_type,
    parse_bool,
    parse_decimal,
    parse_int,
    split_employee_phones,
    split_multiline,
)

DEFAULT_FILE = os.path.join("docs", "МАЛИКА АДМИНСТРАЦИЯГА.xlsx")

# ============================================================
# BIRINCHI SHEET (index 0) — 40 ustun, to'liq ma'lumot
# ============================================================
COL_BLOCK          = 1
COL_SHOP_NUM       = 2
COL_OWNER_FIO      = 3
COL_OWNER_JSHSHIR  = 4
COL_OWNER_PHONE    = 5
COL_OWNER_COMPANY  = 6
COL_TOTAL_AREA     = 7
COL_TENANTS_COUNT  = 8
COL_RENTED_AREA    = 9
COL_TENANT_NAME    = 10
COL_LEADER_JSHSHIR = 11
COL_LEADER_FIO     = 12
COL_LEADER_PHONE   = 13
COL_YTT_JSHSHIR    = 14
COL_STIR           = 15
COL_EMP_COUNT      = 16
COL_EMP_FIO        = 17
COL_EMP_JSHSHIR    = 18
COL_EMP_PHONE      = 19
COL_BUSINESS_TYPE  = 20
COL_TAX_TYPE       = 21
COL_CASH_REG       = 22
COL_MONTHLY_VISIT  = 23
COL_DAILY_VISIT    = 24
COL_ACTIVITY       = 25
COL_FIRE_SAFETY    = 26
COL_FIRE_ALARM     = 27
COL_EXTINGUISHER   = 28
COL_RED_REASON     = 29
COL_YTD_OKKM       = 30
COL_MTD_OKKM       = 32
COL_MTD_OKKM_CHK   = 33
COL_DTD_OKKM       = 34
COL_DTD_OKKM_CHK   = 35
COL_YTD_EPAY_SUM   = 37
COL_MTD_EPAY_SUM   = 39


def _trunc(value, max_len):
    if value is None:
        return None
    return str(value)[:max_len]


def _get(row, col):
    if col >= len(row):
        return None
    return row[col]


class Command(BaseCommand):
    help = (
        "docs/МАЛИКА АДМИНСТРАЦИЯГА.xlsx faylidan do'kon, ijrachi va "
        "xodimlarni import qiladi. Egasi almashgan bo'lsa oxirgi "
        "bo'sh bo'lmagan qator egasi olinadi."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=None,
            help=f"Excel fayl yo'li (default: {DEFAULT_FILE})",
        )
        parser.add_argument(
            "--sheet",
            default=0,
            type=int,
            help="Excel sheet indeksi 0-dan (default: 0 — birinchi sheet, to'liq ma'lumot)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Ma'lumotlar bazasiga yozmasdan natijani ko'rsatadi",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        if not file_path:
            file_path = os.path.join(settings.BASE_DIR, DEFAULT_FILE)

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Fayl topilmadi: {file_path}"))
            return

        sheet_index = options["sheet"]
        dry_run = options["dry_run"]

        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames

        if sheet_index >= len(sheet_names):
            self.stderr.write(
                self.style.ERROR(
                    f"Sheet {sheet_index} topilmadi. Mavjud: {sheet_names}"
                )
            )
            return

        ws = wb[sheet_names[sheet_index]]
        self.stdout.write(
            f"Fayl  : {file_path}\n"
            f"Sheet : [{sheet_index}] \"{sheet_names[sheet_index]}\" "
            f"({ws.max_row} qator x {ws.max_column} ustun)"
        )

        # 1-PASS: barcha qatorlarni (block_type, shop_number) bo'yicha guruhlash
        shop_rows: OrderedDict = OrderedDict()
        skipped_rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_num = _get(row, 0)
            if raw_num is None:
                continue
            try:
                int(str(raw_num).strip())
            except (TypeError, ValueError):
                continue  # header yoki bo'sh qator

            block_type = normalize_block_type(_get(row, COL_BLOCK))
            shop_number = normalize_shop_number(_get(row, COL_SHOP_NUM))

            if not block_type or not shop_number:
                skipped_rows += 1
                continue

            key = (block_type, shop_number)
            if key not in shop_rows:
                shop_rows[key] = []
            shop_rows[key].append(row)

        self.stdout.write(
            f"\n{len(shop_rows)} ta noyob do'kon topildi. "
            f"{skipped_rows} qator blok/raqam yo'qligi sababli o'tkazildi."
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("--- DRY RUN rejimi ---"))

        created_shops = 0
        updated_shops = 0
        created_tenants = 0
        updated_tenants = 0
        created_employees = 0
        error_list = []

        # 2-PASS: har bir do'kon uchun oxirgi egasi + barcha tenantlar
        for (block_type, shop_number), rows in shop_rows.items():
            try:
                with transaction.atomic():
                    # Egasi: oxirgi bo'sh bo'lmagan qator (egasi almashgan holat)
                    owner_fio = owner_jshshir = owner_phone = owner_company = None
                    for row in reversed(rows):
                        candidate = empty_or_zero_to_none(_get(row, COL_OWNER_FIO))
                        if candidate:
                            owner_fio     = _trunc(candidate, 255)
                            owner_jshshir = _trunc(normalize_jshshir(_get(row, COL_OWNER_JSHSHIR)), 255)
                            owner_phone   = _trunc(normalize_phone(_get(row, COL_OWNER_PHONE)), 255)
                            owner_company = _trunc(empty_or_zero_to_none(_get(row, COL_OWNER_COMPANY)), 255)
                            break

                    # Umumiy maydon: birinchi noldan farqli qiymat
                    total_area = None
                    for row in rows:
                        area = parse_decimal(_get(row, COL_TOTAL_AREA))
                        if area and area > 0:
                            total_area = area
                            break

                    shop_data = {
                        "owner_fio":          owner_fio,
                        "owner_jshshir":      owner_jshshir,
                        "owner_phone":        owner_phone,
                        "owner_company_name": owner_company,
                        "total_area":         total_area,
                        "tenants_count":      len(rows),
                    }

                    if not dry_run:
                        shop, shop_created = Shop.objects.update_or_create(
                            block_type=block_type,
                            shop_number=shop_number,
                            defaults=shop_data,
                        )
                        created_shops += 1 if shop_created else 0
                        updated_shops  += 0 if shop_created else 1
                    else:
                        shop = None
                        created_shops += 1

                    # Har bir tenant qatori
                    for row in rows:
                        tenant_name = _trunc(empty_or_zero_to_none(_get(row, COL_TENANT_NAME)), 255)

                        leader_jshshir = normalize_jshshir(_get(row, COL_LEADER_JSHSHIR))
                        if not leader_jshshir:
                            leader_jshshir = normalize_jshshir(_get(row, COL_YTT_JSHSHIR))
                        leader_jshshir = _trunc(leader_jshshir, 255)

                        leader_fio   = _trunc(empty_or_zero_to_none(_get(row, COL_LEADER_FIO)), 255)
                        leader_phone = _trunc(normalize_phone(_get(row, COL_LEADER_PHONE)), 255)
                        stir         = _trunc(normalize_jshshir(_get(row, COL_STIR)), 30)

                        if not tenant_name and not leader_fio and not leader_jshshir:
                            continue

                        tax_type = normalize_tax_type(_get(row, COL_TAX_TYPE))
                        is_vat   = tax_type == ShopTenant.TaxType.VAT
                        cash_reg = _trunc(empty_or_zero_to_none(_get(row, COL_CASH_REG)), 80)

                        red_reason   = _trunc(empty_or_zero_to_none(_get(row, COL_RED_REASON)), 255)
                        extinguisher = _trunc(empty_or_zero_to_none(_get(row, COL_EXTINGUISHER)), 255)

                        ytd_okkm = parse_decimal(_get(row, COL_YTD_OKKM))
                        mtd_okkm = parse_decimal(_get(row, COL_MTD_OKKM))
                        dtd_okkm = parse_decimal(_get(row, COL_DTD_OKKM))
                        ytd_epay = parse_decimal(_get(row, COL_YTD_EPAY_SUM))
                        mtd_epay = parse_decimal(_get(row, COL_MTD_EPAY_SUM))
                        mtd_chk  = parse_int(_get(row, COL_MTD_OKKM_CHK))
                        dtd_chk  = parse_int(_get(row, COL_DTD_OKKM_CHK))

                        tenant_data = {
                            "name":            tenant_name,
                            "leader_fio":      leader_fio,
                            "leader_jshshir":  leader_jshshir,
                            "leader_phone":    leader_phone,
                            "stir":            stir,
                            "certificate_number": None,
                            "rented_area":     parse_decimal(_get(row, COL_RENTED_AREA)),
                            "business_type":   normalize_business_type(_get(row, COL_BUSINESS_TYPE)),
                            "tax_type":        tax_type,
                            "activity_status": normalize_activity_status(_get(row, COL_ACTIVITY)),
                            "employees_count": parse_int(_get(row, COL_EMP_COUNT)),
                            # Kassa apparat
                            "cash_register_number_vat":      cash_reg if is_vat else None,
                            "cash_register_number_turnover": None if is_vat else cash_reg,
                            # Tushum — OKKM
                            "ytd_okkm_vat":      ytd_okkm if is_vat else None,
                            "ytd_okkm_turnover": None if is_vat else ytd_okkm,
                            "mtd_okkm_vat":      mtd_okkm if is_vat else None,
                            "mtd_okkm_turnover": None if is_vat else mtd_okkm,
                            "dtd_okkm_vat":      dtd_okkm if is_vat else None,
                            "dtd_okkm_turnover": None if is_vat else dtd_okkm,
                            # Tushum — Elektron to'lov
                            "ytd_e_payment_vat":      ytd_epay if is_vat else None,
                            "ytd_e_payment_turnover": None if is_vat else ytd_epay,
                            "mtd_e_payment_vat":      mtd_epay if is_vat else None,
                            "mtd_e_payment_turnover": None if is_vat else mtd_epay,
                            "dtd_e_payment_vat":      None,
                            "dtd_e_payment_turnover": None,
                            # Cheklar soni
                            "monthly_checks_count_vat":      mtd_chk if is_vat else None,
                            "monthly_checks_count_turnover": None if is_vat else mtd_chk,
                            "daily_checks_count_vat":        dtd_chk if is_vat else None,
                            "daily_checks_count_turnover":   None if is_vat else dtd_chk,
                            # Tashrif
                            "monthly_visitors": parse_int(_get(row, COL_MONTHLY_VISIT)),
                            "daily_visitors":   parse_int(_get(row, COL_DAILY_VISIT)),
                            # Yong'in xavfsizligi
                            "fire_safety_level": normalize_fire_safety_level(_get(row, COL_FIRE_SAFETY)),
                            "has_fire_alarm":    parse_bool(_get(row, COL_FIRE_ALARM)),
                            "extinguisher_info": extinguisher,
                            # Qizil toifa
                            "is_red_category": bool(red_reason),
                            "red_reason":      red_reason,
                        }

                        if not dry_run:
                            tenant = None
                            if leader_jshshir:
                                tenant = ShopTenant.objects.filter(
                                    shop=shop, leader_jshshir=leader_jshshir
                                ).first()
                            if not tenant and tenant_name:
                                tenant = ShopTenant.objects.filter(
                                    shop=shop, name=tenant_name
                                ).first()

                            if tenant:
                                for field, val in tenant_data.items():
                                    setattr(tenant, field, val)
                                tenant.save()
                                updated_tenants += 1
                            else:
                                tenant = ShopTenant.objects.create(
                                    shop=shop, **tenant_data
                                )
                                created_tenants += 1

                            # Xodimlar (har safar qaytadan yoziladi)
                            emp_fios     = split_multiline(_get(row, COL_EMP_FIO))
                            emp_jshshirs = split_multiline(_get(row, COL_EMP_JSHSHIR))
                            emp_phones   = split_employee_phones(_get(row, COL_EMP_PHONE))

                            tenant.employees.all().delete()

                            max_emp = max(len(emp_fios), len(emp_jshshirs), len(emp_phones))
                            for i in range(max_emp):
                                fio     = _trunc(empty_or_zero_to_none(emp_fios[i] if i < len(emp_fios) else None), 255)
                                jshshir = _trunc(normalize_jshshir(emp_jshshirs[i] if i < len(emp_jshshirs) else None), 20)
                                phone   = _trunc(normalize_phone(emp_phones[i] if i < len(emp_phones) else None), 30)

                                if not fio and not jshshir and not phone:
                                    continue

                                TenantEmployee.objects.create(
                                    tenant=tenant, fio=fio, jshshir=jshshir, phone=phone
                                )
                                created_employees += 1
                        else:
                            created_tenants += 1

            except Exception as exc:
                label = f"Blok {block_type} - {shop_number}-do'kon"
                error_list.append(f"{label}: {exc}")
                self.stderr.write(self.style.WARNING(f"Xato [{label}]: {exc}"))

        prefix = "(DRY RUN) " if dry_run else ""
        self.stdout.write(
            self.style.SUCCESS(
                f"\n{prefix}Import yakunlandi:\n"
                f"  Do'konlar  : {created_shops} yangi, {updated_shops} yangilandi\n"
                f"  Ijarachilar: {created_tenants} yangi, {updated_tenants} yangilandi\n"
                f"  Xodimlar   : {created_employees} yangi\n"
                f"  Xatolar    : {len(error_list)} ta"
            )
        )

        if error_list:
            self.stdout.write("\nXatolar ro'yxati:")
            for err in error_list:
                self.stdout.write(f"  - {err}")
