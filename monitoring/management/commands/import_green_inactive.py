"""
xlsx yashil (theme6) YTT'lar — ketgan tadbirkorlar (o'rniga yangi odam kelgan).
Ularni backendda YO'Q bo'lsa NOFAOL (INACTIVE) ShopTenant qilib qo'shadi va
hozirgi kungacha barcha tushum/soliqini hisoblab (backfill) qo'yadi:

  - ShopTenant: YTT, aylanma soliq, activity_status=INACTIVE, leader_jshshir=PINFL,
    stir=soliqdan resolved tin (bo'lsa), kassa -> cash_register_number_turnover,
    do'kon -> Shop (block+raqam) get_or_create.
  - Faktura : sellerTin = key -> FacturaRevenueDaily(seller_tin=key)
  - OKKM    : kassa terminal + key -> OkkmRevenueDaily(tin=key)
  - Tax     : key (resolved tin) -> TaxRevenue(tin=key, tax_code)  [stir resolve bo'lsa]
    (key = resolved tin bo'lsa o'sha, aks holda PINFL)

Nofaol tenant tax'i sync_tax_revenue da O'CHMAYDI (clear endi barcha tenant
stir/PINFL ni saqlaydi). Faktura/OKKM kunlik sync'da saqlanadi (nofaol skip).

DIQQAT: soliq serverida ishlating. Default DRY-RUN; --apply bilan yozadi. OKKM sekin.

    python manage.py import_green_inactive                 # dry-run
    python manage.py import_green_inactive --apply
    python manage.py import_green_inactive --apply --from 01.01.2026 --to 21.06.2026
    python manage.py import_green_inactive --apply --no-tax     # tax'siz
"""
import os
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

import requests
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from monitoring.models import (
    FacturaRevenueDaily, OkkmRevenueDaily, ShopTenant, Shop, TaxRevenue,
)
from monitoring.services import soliq_service
from monitoring.services.tax_revenue_sync import TAX_CODES

DEFAULT_XLSX = os.path.join("docs", "shop_tenants_2026-06-19.xlsx")

_BLOCK_LETTER = {
    "A": Shop.BlockType.BLOK_A, "B": Shop.BlockType.BLOK_B,
    "G": Shop.BlockType.BLOK_G, "J": Shop.BlockType.BLOK_J,
}


def _digits(v):
    return re.sub(r"\D", "", str(v or ""))


def _to_decimal(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(0)


def _daterange(start, end):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def _parse_shop(dokon):
    """'Blok A - 2' -> (BLOK_A, '2'); 'Merkato' -> (SAVDO_MARKAZ, 'Merkato')."""
    s = str(dokon or "").strip()
    if not s:
        return Shop.BlockType.SAVDO_MARKAZ, "Merkato"
    if "merkato" in s.lower():
        return Shop.BlockType.SAVDO_MARKAZ, "Merkato"
    m = re.search(r"blok\s*([a-zA-Z])\s*-?\s*(\d+)", s, re.I)
    if m:
        block = _BLOCK_LETTER.get(m.group(1).upper())
        if block:
            return block, m.group(2)
    return Shop.BlockType.SAVDO_MARKAZ, s[:50]


def _ckey(cell):
    f = cell.fill
    if not f or not f.patternType:
        return None
    fg = f.fgColor
    if fg.type == "rgb":
        rgb = str(fg.rgb)
        return None if rgb in ("00000000", "FFFFFFFF") else f"rgb:{rgb}"
    if fg.type == "theme":
        return f"theme{fg.theme}/{round(float(fg.tint or 0), 2)}"
    return None


def _is_green(ws, r):
    return any((_ckey(ws.cell(r, c)) or "").startswith("theme6") for c in range(1, 13))


class Command(BaseCommand):
    help = "Yashil YTT'lar (ketgan) ni NOFAOL tenant qilib qo'shadi + tushum/soliqini backfill qiladi."

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", default=None)
        parser.add_argument("--apply", action="store_true", help="Haqiqatan yozadi (default: dry-run)")
        parser.add_argument("--from", dest="period_from", help="dd.mm.yyyy (default: 01.01.joriy_yil)")
        parser.add_argument("--to", dest="period_to", help="dd.mm.yyyy (default: bugun)")
        parser.add_argument("--no-tax", action="store_true", help="Tax-revenue backfill qilinmaydi")
        parser.add_argument("--no-okkm", action="store_true", help="OKKM backfill qilinmaydi (tez)")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        path = opts["xlsx"] or os.path.join(settings.BASE_DIR, DEFAULT_XLSX)
        if not os.path.exists(path):
            raise CommandError(f"Fayl topilmadi: {path}")

        apply = opts["apply"]
        do_tax = not opts["no_tax"]
        do_okkm = not opts["no_okkm"]
        today = date.today()
        pf = opts["period_from"] or f"01.01.{today.year}"
        pt = opts["period_to"] or today.strftime("%d.%m.%Y")
        start = datetime.strptime(pf, "%d.%m.%Y").date()
        end = datetime.strptime(pt, "%d.%m.%Y").date()

        wb = load_workbook(path)
        ws = wb["Shop Tenants"] if "Shop Tenants" in wb.sheetnames else wb.active

        rows = []
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is None or not _is_green(ws, r):
                continue
            pinfl = _digits(ws.cell(r, 11).value)
            if not pinfl:
                continue
            rows.append({
                "pinfl": pinfl,
                "name": (str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else None),
                "dokon": ws.cell(r, 2).value,
                "terminals": soliq_service.parse_terminal_ids(ws.cell(r, 18).value, ws.cell(r, 19).value),
            })

        existing = set(_digits(p) for p in ShopTenant.objects
                       .exclude(leader_jshshir__isnull=True).exclude(leader_jshshir="")
                       .values_list("leader_jshshir", flat=True))
        todo = [g for g in rows if g["pinfl"] not in existing]

        self.stdout.write(f"Fayl: {path}\nDavr: {pf} .. {pt}")
        self.stdout.write(f"Yashil jami: {len(rows)} | backendda yo'q (qo'shiladi): {len(todo)}")
        self.stdout.write(self.style.SUCCESS("--- APPLY ---") if apply else self.style.WARNING("--- DRY-RUN ---"))

        created = fac_rows = okkm_rows = tax_rows = 0
        for g in todo:
            pinfl = g["pinfl"]
            # 1) soliqdan resolved tin (stir) + nom
            tin = None
            try:
                se = soliq_service.get_self_employment(pinfl)
                if se and se.get("tin"):
                    tin = str(se["tin"]).strip()
            except (soliq_service.SoliqError, requests.RequestException):
                pass
            key = tin or pinfl  # faktura/okkm/tax kaliti

            block, num = _parse_shop(g["dokon"])
            kassa = ", ".join(g["terminals"]) if g["terminals"] else None

            self.stdout.write(f"  + {g['name']} | pinfl={pinfl} tin={tin or '-'} | {g['dokon']} | kassa={kassa or '-'}")

            if apply:
                with transaction.atomic():
                    shop, _ = Shop.objects.get_or_create(block_type=block, shop_number=num)
                    ShopTenant.objects.create(
                        shop=shop,
                        name=g["name"],
                        leader_jshshir=pinfl,
                        stir=tin,
                        business_type=ShopTenant.BusinessType.YTT,
                        tax_type=ShopTenant.TaxType.MONTHLY_INCOME,
                        activity_status=ShopTenant.ActivityStatus.INACTIVE,
                        cash_register_number_turnover=kassa,
                    )
            created += 1

            # 2) Faktura backfill (sellerTin = key)
            try:
                daily = soliq_service.aggregate_factura_by_day(key, pf, pt)
            except (soliq_service.SoliqError, requests.RequestException) as e:
                self.stdout.write(self.style.WARNING(f"     faktura xato: {e}"))
                daily = {}
            if apply and daily:
                with transaction.atomic():
                    for day, agg in daily.items():
                        FacturaRevenueDaily.objects.update_or_create(
                            date=day, seller_tin=key,
                            defaults={"sales": agg["sales"], "tax": agg["tax"], "factura_count": agg["count"]},
                        )
            fac_rows += len(daily)

            # 3) OKKM backfill (kassa terminal + key)
            if do_okkm and g["terminals"]:
                okkm_rows += self._okkm(key, g["terminals"], start, end, apply)

            # 4) Tax backfill (faqat resolved tin bo'lsa)
            if do_tax and tin:
                tax_rows += self._tax(key, today, apply)

        self.stdout.write(self.style.SUCCESS(
            f"\n{'Yozildi' if apply else 'Yoziladi'}: tenant={created}, "
            f"factura_kun={fac_rows}, okkm_kun={okkm_rows}, tax_yozuv={tax_rows}"
        ))

    # ------------------------------------------------------------------
    def _okkm(self, key, terminals, start, end, apply):
        written = 0
        for day in _daterange(start, end):
            day_str = day.strftime("%Y-%m-%d")
            count = 0
            turnover = vat = Decimal(0)
            for terminal in terminals:
                try:
                    s = soliq_service.get_cheque_summary(key, terminal, day_str, day_str)
                except (requests.RequestException, ValueError):
                    continue
                if not s:
                    continue
                count += int(s.get("chequeCount") or 0)
                turnover += _to_decimal(s.get("turnover"))
                vat += _to_decimal(s.get("vat"))
            if turnover == 0 and vat == 0 and count == 0:
                continue
            if apply:
                with transaction.atomic():
                    OkkmRevenueDaily.objects.update_or_create(
                        date=day, tin=key,
                        defaults={"turnover": turnover, "vat": vat, "cheque_count": count},
                    )
            written += 1
        return written

    # ------------------------------------------------------------------
    def _tax(self, tin, today, apply):
        ytd_from = f"01.01.{today.year}"
        mtd_from = today.strftime("01.%m.%Y")
        today_str = today.strftime("%d.%m.%Y")
        written = 0
        for code, _name in TAX_CODES:
            try:
                ytd = _to_decimal((soliq_service.get_company_account(tin, ytd_from, today_str, code) or {}).get("payTax"))
                mtd = _to_decimal((soliq_service.get_company_account(tin, mtd_from, today_str, code) or {}).get("payTax"))
                dtd = _to_decimal((soliq_service.get_company_account(tin, today_str, today_str, code) or {}).get("payTax"))
            except (soliq_service.SoliqError, requests.RequestException):
                continue
            if apply:
                with transaction.atomic():
                    TaxRevenue.objects.update_or_create(
                        tin=tin, tax_code=code,
                        defaults={"dtd_pay": dtd, "mtd_pay": mtd, "ytd_pay": ytd},
                    )
            written += 1
        return written
