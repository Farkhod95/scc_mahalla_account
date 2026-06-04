"""
malumot.xlsx dan kadastr raqamini do'konlarga (Shop) yozadi.

Ustunlar: Blok, Do'kon raqami, Kadastr raqami (D ustun).
Har do'kon (Blok + raqam) bo'yicha kadastr yoziladi (tenant qatorlari takror,
shuning uchun do'kon bo'yicha dedup qilinadi).

Ishlatish:
    python manage.py import_shop_cadastr
    python manage.py import_shop_cadastr --file "docs/malumot.xlsx" --dry-run
"""
import os

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from monitoring.models import Shop
from monitoring.tasks import normalize_block_type, normalize_shop_number, clean_single_line

DEFAULT_FILE = os.path.join("docs", "malumot.xlsx")


def _find_columns(ws):
    """Header qatorini va block/shop_number/cadastr ustun indekslarini topadi."""
    for r in range(1, 8):
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            t = v.strip().lower()
            if t == "blok":
                cols["block"] = c
            elif "do'kon raqami" in t or "dokon raqami" in t or "do‘kon raqami" in t:
                cols["shop_number"] = c
            elif "kadastr" in t:
                cols["cadastr"] = c
        if "block" in cols and "cadastr" in cols and "shop_number" in cols:
            return r, cols
    return None, {}


class Command(BaseCommand):
    help = "Excel'dan kadastr raqamini do'konlarga yozadi."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=None, help=f"Excel (default: {DEFAULT_FILE})")
        parser.add_argument("--dry-run", action="store_true", help="DBga yozmaydi")

    def handle(self, *args, **opts):
        file_path = opts["file"] or os.path.join(settings.BASE_DIR, DEFAULT_FILE)
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Fayl topilmadi: {file_path}"))
            return

        wb = load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        header_row, cols = _find_columns(ws)
        if not header_row:
            self.stderr.write(self.style.ERROR("Blok / Do'kon raqami / Kadastr ustunlari topilmadi"))
            return
        self.stdout.write(f"Ustunlar: {cols} (header qator {header_row})")

        updated = skipped = not_found = 0
        seen = set()

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            block_raw = row[cols["block"] - 1]
            num_raw = row[cols["shop_number"] - 1]
            cadastr = clean_single_line(row[cols["cadastr"] - 1])

            block = normalize_block_type(block_raw)
            shop_number = normalize_shop_number(num_raw)

            if not block or not shop_number or not cadastr:
                skipped += 1
                continue

            key = (block, shop_number)
            if key in seen:
                continue
            seen.add(key)

            shop = Shop.objects.filter(block_type=block, shop_number=shop_number).first()
            if not shop:
                not_found += 1
                self.stdout.write(f"  [topilmadi] {block}-{shop_number} | {cadastr}")
                continue

            self.stdout.write(f"  {block}-{shop_number} -> {cadastr}")
            if not opts["dry_run"]:
                shop.cadastr_number = cadastr
                shop.save(update_fields=["cadastr_number"])
            updated += 1

        prefix = "(DRY RUN) " if opts["dry_run"] else ""
        self.stdout.write(self.style.SUCCESS(
            f"\n{prefix}Tugadi: yozildi={updated}, do'kon topilmadi={not_found}, o'tkazildi={skipped}"
        ))
