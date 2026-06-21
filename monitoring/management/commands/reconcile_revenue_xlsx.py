"""
shop_tenants_*.xlsx (qo'lda belgilangan) bo'yicha savdo tushumini tuzatadi:

  - SARIQ (rgb FFFF00) qatorlar = tushumi XATO/takror tashkilotlar. Tashkilot
    (ShopTenant) O'CHIRILMAYDI — faqat unga ulangan tushum datasi (FacturaRevenueDaily,
    ixtiyoriy OkkmRevenueDaily) tin bo'yicha o'chiriladi.
        tin = STIR (bo'lsa), aks holda PINFL (leader JSHSHIR) -> soliq self-employment tin.
  - YASHIL (theme6, och yashil) qatorlar = yangi YTT'lar. Ularning tushumi PINFL
    bo'yicha soliqdan qaytadan hisoblanadi (tenant_soliq_sync._sync_one).

Sana/rang faqat birinchi 12 ustun (ID..izoh) bo'yicha aniqlanadi.

DIQQAT: soliqqa boradi (YTT tin resolve + green re-sync) — soliq ochiladigan
SERVERDA ishlating. Default DRY-RUN; haqiqiy o'zgarish uchun --apply.

    python manage.py reconcile_revenue_xlsx                       # dry-run (default fayl)
    python manage.py reconcile_revenue_xlsx --apply               # bajaradi (factura data o'chadi + green re-sync)
    python manage.py reconcile_revenue_xlsx --apply --okkm        # OkkmRevenueDaily ni ham o'chiradi
    python manage.py reconcile_revenue_xlsx --yellow-only --apply # faqat yellow tozalash
    python manage.py reconcile_revenue_xlsx --green-only --apply  # faqat green re-sync
    python manage.py reconcile_revenue_xlsx --xlsx /path/to.xlsx
"""
import os
import re

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from monitoring.models import FacturaRevenueDaily, OkkmRevenueDaily, ShopTenant
from monitoring.services.revenue_sync import _tin_from_pinfl
from monitoring.services.tenant_soliq_sync import _sync_one

DEFAULT_XLSX = os.path.join("docs", "shop_tenants_2026-06-19.xlsx")

YELLOW = "rgb:FFFFFF00"


def _digits(v):
    return re.sub(r"\D", "", str(v or ""))


def _ckey(cell):
    f = cell.fill
    if not f or not f.patternType:
        return None
    fg = f.fgColor
    if fg.type == "rgb":
        rgb = str(fg.rgb)
        return None if rgb in ("00000000", "FFFFFFFF") else f"rgb:{rgb}"
    if fg.type == "theme":
        return f"theme{fg.theme}/{round(float(fg.tint or 0), 2)}"
    return None


def _row_color(ws, r):
    cols = [_ckey(ws.cell(r, c)) for c in range(1, 13)]
    cols = [c for c in cols if c]
    if YELLOW in cols:
        return "YELLOW"
    if any(c.startswith("theme6") for c in cols):
        return "GREEN"
    return None


class Command(BaseCommand):
    help = "shop_tenants xlsx (sariq/yashil) bo'yicha tushum datasini tuzatadi."

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", default=None, help=f"Excel yo'li (default: {DEFAULT_XLSX})")
        parser.add_argument("--apply", action="store_true", help="Haqiqatan bajaradi (default: dry-run)")
        parser.add_argument("--okkm", action="store_true", help="OkkmRevenueDaily ni ham o'chiradi (default: faqat factura)")
        parser.add_argument("--yellow-only", action="store_true", help="Faqat yellow (factura data o'chirish)")
        parser.add_argument("--green-only", action="store_true", help="Faqat green (re-sync)")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        path = opts["xlsx"] or os.path.join(settings.BASE_DIR, DEFAULT_XLSX)
        if not os.path.exists(path):
            raise CommandError(f"Fayl topilmadi: {path}")

        apply = opts["apply"]
        do_okkm = opts["okkm"]
        yellow_only = opts["yellow_only"]
        green_only = opts["green_only"]
        if yellow_only and green_only:
            raise CommandError("--yellow-only va --green-only birga berilmaydi.")

        wb = load_workbook(path)
        ws = wb["Shop Tenants"] if "Shop Tenants" in wb.sheetnames else wb.active

        yellow, green = [], []
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is None:
                continue
            color = _row_color(ws, r)
            if not color:
                continue
            rec = {
                "id": ws.cell(r, 1).value,
                "dokon": ws.cell(r, 2).value,
                "name": ws.cell(r, 3).value,
                "stir": _digits(ws.cell(r, 8).value),
                "pinfl": _digits(ws.cell(r, 11).value),
            }
            (yellow if color == "YELLOW" else green).append(rec)

        self.stdout.write(f"Fayl: {path}")
        self.stdout.write(f"Sariq (yellow): {len(yellow)} qator | Yashil (green): {len(green)} qator")
        self.stdout.write(self.style.WARNING("--- DRY-RUN (--apply yo'q) ---") if not apply else self.style.SUCCESS("--- APPLY ---"))

        if not green_only:
            self._do_yellow(yellow, apply, do_okkm)
        if not yellow_only:
            self._do_green(green, apply)

    # ----------------------------------------------------------------
    def _do_yellow(self, yellow, apply, do_okkm):
        self.stdout.write("\n=== YELLOW: tushum datasini o'chirish ===")
        tins, unresolved = set(), []
        for rec in yellow:
            tin = rec["stir"]
            if not tin and rec["pinfl"]:
                # YTT: pinfl -> soliq self-employment tin. Tenant'ni topib _tin_from_pinfl.
                t = ShopTenant.objects.filter(leader_jshshir=rec["pinfl"]).first()
                if t:
                    tin = _tin_from_pinfl(t)
                if not tin:
                    # tenant yo'q bo'lsa ham pinfl bo'yicha to'g'ridan-to'g'ri soliqqa urinib ko'ramiz
                    tmp = ShopTenant(leader_jshshir=rec["pinfl"])
                    tin = _tin_from_pinfl(tmp)
            if tin:
                tins.add(tin)
            else:
                unresolved.append(rec)

        tins = sorted(tins)
        self.stdout.write(f"Aniqlangan tin: {len(tins)} ta -> {tins}")
        if unresolved:
            self.stdout.write(self.style.WARNING(f"Tin aniqlanmadi: {len(unresolved)} ta (factura datasi bo'lmasligi mumkin):"))
            for rec in unresolved:
                self.stdout.write(f"   {rec['name']} | pinfl={rec['pinfl']} | {rec['dokon']}")

        if not tins:
            return

        f_count = FacturaRevenueDaily.objects.filter(seller_tin__in=tins).count()
        o_count = OkkmRevenueDaily.objects.filter(tin__in=tins).count() if do_okkm else 0
        self.stdout.write(f"O'chiriladi: FacturaRevenueDaily={f_count}" + (f", OkkmRevenueDaily={o_count}" if do_okkm else ""))

        if apply:
            f_del = FacturaRevenueDaily.objects.filter(seller_tin__in=tins).delete()[0]
            msg = f"O'chirildi: factura={f_del}"
            if do_okkm:
                o_del = OkkmRevenueDaily.objects.filter(tin__in=tins).delete()[0]
                msg += f", okkm={o_del}"
            self.stdout.write(self.style.SUCCESS(msg))

    # ----------------------------------------------------------------
    def _do_green(self, green, apply):
        self.stdout.write("\n=== GREEN: tushumni PINFL bo'yicha re-sync ===")
        found, missing, updated = 0, [], 0
        for rec in green:
            pinfl = rec["pinfl"]
            tenants = list(ShopTenant.objects.filter(leader_jshshir=pinfl))
            if not tenants:
                missing.append(rec)
                continue
            found += len(tenants)
            if apply:
                for t in tenants:
                    try:
                        if _sync_one(t):
                            updated += 1
                    except Exception as e:
                        self.stdout.write(self.style.WARNING(f"   re-sync xato (tenant={t.pk}, pinfl={pinfl}): {e}"))

        self.stdout.write(f"Topilgan tenant: {found} ta")
        if missing:
            self.stdout.write(self.style.WARNING(f"Backendda topilmadi: {len(missing)} ta pinfl:"))
            for rec in missing:
                self.stdout.write(f"   {rec['name']} | pinfl={rec['pinfl']} | {rec['dokon']}")
        if apply:
            self.stdout.write(self.style.SUCCESS(f"Re-sync yangilandi: {updated} tenant"))
