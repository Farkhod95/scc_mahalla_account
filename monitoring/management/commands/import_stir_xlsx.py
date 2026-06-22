"""
Faktura xlsx (ustunlar: tenant_id, stir, pinfl, name) dan STIR'larni ShopTenant ga
yozadi. Mos kelish: tenant_id (bizning missing_revenue_list eksporti id si) bo'yicha;
pinfl bo'lsa qo'shimcha tekshiriladi (noto'g'ri tenantga yozmaslik uchun).

  - default: faqat BO'SH stir to'ldiriladi.
  - --overwrite: mavjud (farqli) stir ham almashtiriladi.
  - default DRY-RUN; --apply bilan yozadi.

Stir to'g'rilangach faktura/OKKM ni qaytadan sync qiling (ular stir bo'yicha oladi).

    python manage.py import_stir_xlsx --file "docs/faktur.xlsx"
    python manage.py import_stir_xlsx --file "docs/faktur.xlsx" --apply
    python manage.py import_stir_xlsx --file "docs/faktur.xlsx" --apply --overwrite
"""
import os
import re

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from monitoring.models import ShopTenant

DEFAULT_FILE = os.path.join("docs", "faktur.xlsx")


def _digits(v):
    return re.sub(r"\D", "", str(v or ""))


class Command(BaseCommand):
    help = "Faktura xlsx dan STIR'larni ShopTenant ga yozadi (tenant_id bo'yicha)."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=None, help=f"Excel yo'li (default: {DEFAULT_FILE})")
        parser.add_argument("--apply", action="store_true", help="Haqiqatan yozadi (default: dry-run)")
        parser.add_argument("--overwrite", action="store_true",
                            help="Mavjud (farqli) stir ni ham almashtiradi (default: faqat bo'shini to'ldiradi)")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        path = opts["file"] or os.path.join(settings.BASE_DIR, DEFAULT_FILE)
        if not os.path.exists(path):
            raise CommandError(f"Fayl topilmadi: {path}")

        apply = opts["apply"]
        overwrite = opts["overwrite"]

        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hdr = {str(ws.cell(1, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}
        for req in ("tenant_id", "stir"):
            if req not in hdr:
                raise CommandError(f"'{req}' ustuni topilmadi. Ustunlar: {list(hdr)}")
        ci_id = hdr["tenant_id"]
        ci_stir = hdr["stir"]
        ci_pinfl = hdr.get("pinfl")

        filled = changed = already = no_stir = 0
        updated, notfound, mismatch, conflict = [], [], [], []

        for r in range(2, ws.max_row + 1):
            stir = _digits(ws.cell(r, ci_stir).value)
            if not stir:
                no_stir += 1
                continue
            tid = _digits(ws.cell(r, ci_id).value)
            if not tid:
                notfound.append((None, stir, "tenant_id yo'q"))
                continue
            tenant = ShopTenant.objects.filter(id=int(tid)).first()
            if not tenant:
                notfound.append((tid, stir, "tenant topilmadi"))
                continue

            # pinfl tekshiruvi (noto'g'ri tenantga yozmaslik uchun)
            x_pinfl = _digits(ws.cell(r, ci_pinfl).value) if ci_pinfl else ""
            t_pinfl = _digits(tenant.leader_jshshir)
            if x_pinfl and t_pinfl and x_pinfl != t_pinfl:
                mismatch.append((tid, tenant.name, t_pinfl, x_pinfl))
                continue

            cur = _digits(tenant.stir)
            if cur == stir:
                already += 1
                continue
            if cur and not overwrite:
                conflict.append((tid, tenant.name, cur, stir))
                continue

            if apply:
                tenant.stir = stir
                with transaction.atomic():
                    tenant.save(update_fields=["stir", "updated_time"])
            if cur:
                changed += 1
            else:
                filled += 1
            updated.append((tid, tenant.name, cur or "-", stir))

        # ---- hisobot ----
        self.stdout.write(f"Fayl: {path}")
        self.stdout.write(self.style.SUCCESS("--- APPLY ---") if apply else self.style.WARNING("--- DRY-RUN ---"))
        self.stdout.write(
            f"STIR yozilgan qatorlar: {len(updated) + already + len(mismatch) + len(conflict) + len(notfound)} | "
            f"{'yozildi' if apply else 'yoziladi'}: {len(updated)} "
            f"(bo'sh->to'ldirildi={filled}, almashtirildi={changed}), allaqachon={already}"
        )
        for tid, name, old, new in updated:
            self.stdout.write(f"  id={tid} {name}: {old} -> {new}")

        if conflict:
            self.stdout.write(self.style.WARNING(
                f"\nMavjud farqli stir ({len(conflict)}) — o'zgartirilmadi (--overwrite kerak):"))
            for tid, name, cur, new in conflict:
                self.stdout.write(f"  id={tid} {name}: bor={cur}, xlsx={new}")
        if mismatch:
            self.stdout.write(self.style.WARNING(f"\nPINFL mos kelmadi ({len(mismatch)}) — o'tkazib yuborildi:"))
            for tid, name, tp, xp in mismatch:
                self.stdout.write(f"  id={tid} {name}: backend_pinfl={tp}, xlsx_pinfl={xp}")
        if notfound:
            self.stdout.write(self.style.WARNING(f"\nTopilmadi ({len(notfound)}):"))
            for row in notfound:
                self.stdout.write(f"  {row}")
