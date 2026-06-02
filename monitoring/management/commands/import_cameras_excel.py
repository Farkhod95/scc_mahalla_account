"""
docs/Malika bozor (2).xlsx dan ShopCamera larni import qiladi va go2rtc ga sinxronlaydi.

Sheet 0 — B blok kameralari   (4 ustun: №, IP, joy, Parol)
Sheet 1 — Perimetr kameralari  (4 ustun: №, IP, joy, Parol)

RTSP URL: rtsp://admin:{password}@{ip}:554/Streaming/Channels/101/

Duplicate oldini olish: IP manzili bo'yicha mavjud kamera qidiriladi.
"""
import os
import re

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from monitoring.models import Shop, ShopCamera
from monitoring.services import go2rtc_service

DEFAULT_FILE  = os.path.join("docs", "Malika bozor (2).xlsx")
DEFAULT_LOGIN = "admin"
RTSP_PORT     = "554"
RTSP_PATH     = "/Streaming/Channels/101/"

BLOCK_MAP = {"A": "A", "B": "B", "G": "G", "J": "J"}

# "B8", "G4", "B44/2", "B43a", "B-15 oldi taraf", "A-28 ichki taraf"
_LOC_RE = re.compile(r"^([A-Ga-g])\s*-?\s*(\d+)", re.IGNORECASE)


def _build_rtsp(ip: str, password: str | None, login: str, rtsp_path: str) -> str:
    cred = f"{login}:{password}@" if password else f"{login}@"
    return f"rtsp://{cred}{ip}:{RTSP_PORT}{rtsp_path}"


def _parse_location(raw: str):
    """
    'B8'              → ('B', '8')
    'B44/2'           → ('B', '44')
    'B43a'            → ('B', '43')
    'B-15 oldi taraf' → ('B', '15')
    'Stoyanka'        → (None, None)
    """
    if not raw:
        return None, None
    m = _LOC_RE.match(raw.strip())
    if not m:
        return None, None
    return BLOCK_MAP.get(m.group(1).upper()), m.group(2)


def _find_shop(block_type, shop_number):
    if not block_type or not shop_number:
        return None
    return Shop.objects.filter(block_type=block_type, shop_number=shop_number).first()


def _find_camera_by_ip(ip: str):
    """Mavjud kamerani IP manzili bo'yicha qidiradi (URL ichida)."""
    return ShopCamera.objects.filter(url__contains=f"@{ip}:").first()


class Command(BaseCommand):
    help = (
        "docs/Malika bozor (2).xlsx dan kameralarni import qilib, "
        "ShopCamera yaratadi/yangilaydi va go2rtc ga sinxronlaydi."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=None,
            help=f"Excel fayl yo'li (default: {DEFAULT_FILE})",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="DBga yozmasdan natijani ko'rsatadi",
        )
        parser.add_argument(
            "--login", default=DEFAULT_LOGIN,
            help=f"RTSP login (default: {DEFAULT_LOGIN})",
        )
        parser.add_argument(
            "--rtsp-path", default=RTSP_PATH,
            help=f"RTSP stream path (default: {RTSP_PATH})",
        )

    def handle(self, *args, **options):
        file_path = options["file"] or os.path.join(settings.BASE_DIR, DEFAULT_FILE)

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Fayl topilmadi: {file_path}"))
            return

        dry_run   = options["dry_run"]
        login     = options["login"]
        rtsp_path = options["rtsp_path"]

        wb = load_workbook(file_path, data_only=True)
        self.stdout.write(f"Fayl  : {file_path}\nSheets: {wb.sheetnames}\n")

        created = updated = skipped = synced = 0
        errors = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            self.stdout.write(f"\n--- Sheet {sheet_idx}: \"{sheet_name}\" "
                              f"({ws.max_row - 1} qator) ---")

            for row in ws.iter_rows(min_row=2, values_only=True):
                num = row[0] if len(row) > 0 else None
                ip  = str(row[1]).strip() if len(row) > 1 and row[1] else None
                loc = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                pwd = str(row[3]).strip() if len(row) > 3 and row[3] else None

                # Bo'sh, sarlavha va NVR qatorlarini o'tkazish
                if not ip:
                    continue
                if loc.upper() == "NVR":
                    skipped += 1
                    continue
                try:
                    int(str(num))
                except (TypeError, ValueError):
                    continue

                rtsp_url   = _build_rtsp(ip, pwd, login, rtsp_path)
                block, num_str = _parse_location(loc)
                shop = _find_shop(block, num_str) if not dry_run else None

                shop_label = f"{block}-{num_str}" if block else "—"
                self.stdout.write(
                    f"  {ip:18s} | {loc:30s} | {shop_label:8s} | "
                    f"{'✓ ' + str(shop) if shop else '—'}"
                )

                if dry_run:
                    created += 1
                    continue

                try:
                    with transaction.atomic():
                        # IP bo'yicha mavjud kamerani topamiz (duplikat oldini olish)
                        existing = _find_camera_by_ip(ip)

                        if existing:
                            existing.title = loc
                            existing.url   = rtsp_url
                            existing.shop  = shop
                            existing.save()
                            updated += 1
                            cam = existing
                        else:
                            cam = ShopCamera.objects.create(
                                title=loc, url=rtsp_url, shop=shop
                            )
                            created += 1

                        # Signal allaqachon chaqiradi, lekin go2rtc
                        # javobini tekshirib qo'yamiz
                        if go2rtc_service.register_stream(cam):
                            synced += 1

                except Exception as exc:
                    errors.append(f"{ip} | {loc}: {exc}")
                    self.stderr.write(self.style.WARNING(f"  Xato [{ip}]: {exc}"))

        prefix = "(DRY RUN) " if dry_run else ""
        self.stdout.write(self.style.SUCCESS(
            f"\n{prefix}Import yakunlandi:\n"
            f"  Yaratildi : {created}\n"
            f"  Yangilandi: {updated}\n"
            f"  O'tkazildi: {skipped} (NVR)\n"
            f"  go2rtc    : {synced} ta sinxronlandi\n"
            f"  Xatolar   : {len(errors)} ta"
        ))
        for e in errors:
            self.stdout.write(f"  - {e}")
