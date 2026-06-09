from typing import Dict, Any, List

from monitoring.services.citizen_avatar import fetch_citizen_avatar_file
from monitoring.services.patrol_car_sync import sync_patrol_cars_from_api
from monitoring.services.revenue_sync import sync_factura_revenue
import os
import re
from decimal import Decimal, InvalidOperation

from celery import shared_task
from django.db import transaction
from openpyxl import load_workbook

from .models import Shop, ShopTenant, TenantEmployee


# =========================================================
# FIELD LABELS
# =========================================================

FIELD_LABELS = {
    "block": "Blok",
    "shop_number": "Do'kon raqami",
    "shop_owner_full_name": "Do'kon egasi F.I.Sh",
    "shop_owner_jshshir": "Do'kon egasi JSHSHIR",
    "shop_owner_phone": "Do'kon egasi telefon",
    "shop_owner_company_name": "Do'kon egasiga tegishli MCHJ nomi",
    "shop_total_area": "Do'kon umumiy maydoni",
    "tenant_count": "Ijaraga olgan tadbirkor soni",
    "rented_area_sqm": "Ijaraga berilgan maydoni",
    "business_name": "Tadbirkorlik subyekti nomi",
    "business_leader_jshshir": "Rahbar JSHSHIR",
    "business_leader_full_name": "Rahbar F.I.Sh",
    "business_leader_phone": "Rahbar telefon",
    "business_inn": "Guvohnoma/STIR",
    "employee_count": "Xodimlar soni",
    "employee_full_name": "Xodim F.I.Sh",
    "employee_jshshir": "Xodim JSHSHIR",
    "employee_phone": "Xodim telefon",
    "business_type": "Tadbirkorlik turi",
    "tax_type": "Soliq turi",
    "cash_register_number": "Kassa apparat raqami",
    "yearly_revenue_cash_register": "Yillik tushum OKKM",
    "yearly_revenue_invoice": "Yillik tushum EHF",
    "yearly_revenue_qr": "Yillik tushum QR",
    "monthly_revenue_cash_register": "Oylik tushum OKKM",
    "monthly_revenue_invoice": "Oylik tushum EHF",
    "monthly_revenue_qr": "Oylik tushum QR",
    "daily_revenue_cash_register": "Kunlik tushum OKKM",
    "daily_revenue_invoice": "Kunlik tushum EHF",
    "daily_revenue_qr": "Kunlik tushum QR",
    "monthly_receipt_count": "Oylik cheklar soni",
    "daily_receipt_count": "Kunlik cheklar soni",
    "monthly_visitors": "Oylik tashrif buyuruvchilar",
    "daily_visitors": "Kunlik tashrif buyuruvchilar",
    "business_status": "Subyekt holati",
    "fire_safety_level": "Yong'in xavfsizlik darajasi",
    "fire_alarm": "Yong'indan xabar beruvchi qurilma",
    "fire_extinguisher": "Ognetushitel turi/soni",
    "red_category_reason": "Qizil toifa sababi",
}

MODEL_FIELD_MAX_LENGTHS = {
    "Shop.owner_fio": 255,
    "Shop.owner_jshshir": 20,
    "Shop.owner_phone": 30,
    "Shop.owner_company_name": 255,
    "Shop.shop_number": 120,

    "ShopTenant.name": 255,
    "ShopTenant.leader_fio": 255,
    "ShopTenant.leader_jshshir": 20,
    "ShopTenant.leader_phone": 30,
    "ShopTenant.stir": 30,
    "ShopTenant.certificate_number": 80,
    # Bir nechta kassa raqami vergul bilan saqlanadi (masalan parkovkada 4 ta) —
    # modeldagi haqiqiy uzunlik (255) bilan moslashtirildi, aks holda import rad etardi.
    "ShopTenant.cash_register_number_vat": 255,
    "ShopTenant.cash_register_number_turnover": 255,
    "ShopTenant.extinguisher_info": 255,
    "ShopTenant.red_reason": 255,

    "TenantEmployee.fio": 255,
    "TenantEmployee.jshshir": 20,
    "TenantEmployee.phone": 30,
}


# =========================================================
# EXCEPTIONS
# =========================================================

class RowImportError(Exception):
    def __init__(self, field_key, model_field, message, value=None):
        self.field_key = field_key
        self.model_field = model_field
        self.message = message
        self.value = value
        super().__init__(message)


# =========================================================
# CLEAN HELPERS
# =========================================================

def clean_text(value):
    if value is None:
        return None
    value = str(value).replace("\xa0", " ").strip()
    value = value.replace("\r", "\n")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n+", "\n", value).strip()
    value = value.strip(' "\'')
    return value or None


def clean_single_line(value):
    value = clean_text(value)
    if not value:
        return None
    value = value.replace("\n", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value or None


def is_zero_like(value):
    if value is None:
        return True

    s = clean_single_line(value)
    if not s:
        return True

    lowered = s.lower()

    if lowered in {"0", "0.0", "0,0", "-", "--", "yo'q", "yoq", "none", "null", "nan"}:
        return True

    digits_only = re.sub(r"\D", "", s)
    if digits_only and set(digits_only) == {"0"}:
        return True

    return False


def empty_or_zero_to_none(value):
    if is_zero_like(value):
        return None
    return clean_single_line(value)


def normalize_phone(value):
    value = empty_or_zero_to_none(value)
    if not value:
        return None

    value = re.sub(r"[^0-9\+]", "", value)

    if not value:
        return None

    digits = re.sub(r"\D", "", value)
    if digits and set(digits) == {"0"}:
        return None

    return value


def normalize_jshshir(value):
    value = empty_or_zero_to_none(value)
    if not value:
        return None

    digits = re.sub(r"\D", "", value)
    if not digits:
        return None

    if set(digits) == {"0"}:
        return None

    return digits


def parse_decimal(value):
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        try:
            d = Decimal(str(value))
            if d == 0:
                return Decimal("0")
            return d
        except Exception:
            return None

    s = clean_single_line(value)
    if not s:
        return None

    s = s.replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9\.\-]", "", s)

    if s in {"", "-", ".", "-."}:
        return None

    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_int(value):
    d = parse_decimal(value)
    if d is None:
        return None
    try:
        return int(d)
    except Exception:
        return None


def parse_bool(value):
    value = clean_single_line(value)
    if not value:
        return False

    v = value.lower()
    if v in {"ha", "bor", "true", "1", "yes", "y", "mavjud"}:
        return True
    if v in {"yo'q", "yoq", "false", "0", "no", "n", "mavjud emas"}:
        return False
    return False


def split_multiline(value):
    value = clean_text(value)
    if not value:
        return []

    parts = []
    for line in value.split("\n"):
        line = empty_or_zero_to_none(line)
        if line:
            parts.append(line)
    return parts


def split_employee_phones(value):
    value = clean_text(value)
    if not value:
        return []

    text = str(value).replace("\xa0", " ").strip()
    text = text.replace("\n", " ").replace("\r", " ").replace(";", " ").replace(",", " ")
    text = re.sub(r"\s+", " ", text).strip()

    matches = re.findall(r"\+?\d{7,20}", text)
    result = []

    for phone in matches:
        phone = normalize_phone(phone)
        if not phone:
            continue

        digits = re.sub(r"\D", "", phone)
        if digits and set(digits) == {"0"}:
            continue

        result.append(phone)

    return result


def normalize_key(value):
    value = clean_text(value)
    if not value:
        return ""

    value = value.lower()
    value = value.replace("\n", " ")
    value = value.replace("\r", " ")
    value = value.replace("\t", " ")
    value = value.replace("o‘", "o'")
    value = value.replace("g‘", "g'")
    value = value.replace("“", '"').replace("”", '"')
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^a-zA-Z0-9_ʼ'ʻ]+", "_", value)
    value = value.strip("_")
    return value


def normalize_block_type(value):
    value = clean_single_line(value)
    if not value:
        return None

    v = value.lower()
    if "blok a" in v or v == "a":
        return Shop.BlockType.BLOK_A
    if "blok b" in v or v == "b":
        return Shop.BlockType.BLOK_B
    if "blok g" in v or v == "g":
        return Shop.BlockType.BLOK_G
    if "blok j" in v or v == "j":
        return Shop.BlockType.BLOK_J
    if "savdo markaz" in v or v == "sm":
        return Shop.BlockType.SAVDO_MARKAZ
    if "parkovka" in v or "avtoturargoh" in v or v == "p":
        return Shop.BlockType.PARKOVKA
    return None


def normalize_shop_number(value):
    value = clean_single_line(value)
    if not value:
        return None

    value = value.replace('"', "").replace("'", "").strip()
    value = value.replace("–", "-").replace("—", "-")
    value = re.sub(r"\s+", " ", value)

    m = re.search(r"(\d+)", value)
    if m:
        return m.group(1)

    return value


def normalize_business_type(value):
    value = clean_single_line(value)
    if not value:
        return None
    v = value.lower()

    if "ytt" in v:
        return ShopTenant.BusinessType.YTT
    if "yuridik" in v or "mchj" in v or "ooo" in v or "legal" in v:
        return ShopTenant.BusinessType.LEGAL
    return ShopTenant.BusinessType.OTHER


def normalize_tax_type(value):
    value = clean_single_line(value)
    if not value:
        return None
    v = value.lower()

    if "qqs" in v or "vat" in v:
        return ShopTenant.TaxType.VAT
    if "aos" in v or "aylanma" in v or "monthly_income" in v or "daromad" in v:
        return ShopTenant.TaxType.MONTHLY_INCOME
    return ShopTenant.TaxType.OTHER


def normalize_activity_status(value):
    value = clean_single_line(value)
    if not value:
        return None
    v = value.lower()

    if "faol" in v or "active" in v:
        return ShopTenant.ActivityStatus.ACTIVE
    if "nofaol" in v or "inactive" in v:
        return ShopTenant.ActivityStatus.INACTIVE
    return None


def normalize_fire_safety_level(value):
    value = clean_single_line(value)
    if not value:
        return None
    v = value.lower()

    if "past" in v or "low" in v:
        return ShopTenant.FireSafetyLevel.LOW
    if "o'rtacha" in v or "ortacha" in v or "medium" in v:
        return ShopTenant.FireSafetyLevel.MEDIUM
    if "yuqori" in v or "high" in v:
        return ShopTenant.FireSafetyLevel.HIGH
    return None


def validate_max_length(field_key, model_field, value):
    if value is None:
        return value

    max_len = MODEL_FIELD_MAX_LENGTHS.get(model_field)
    if not max_len:
        return value

    if len(str(value)) > max_len:
        raise RowImportError(
            field_key=field_key,
            model_field=model_field,
            message=f"Qiymat uzunligi {max_len} dan oshib ketgan",
            value=value
        )
    return value


def split_certificate_and_stir(raw_value):
    raw = clean_single_line(raw_value)
    if not raw:
        return None, None

    only_digits = re.sub(r"\D", "", raw)
    if len(only_digits) == 9:
        return only_digits, None

    return None, raw


HEADER_ALIASES = {
    "block": ["block", "blok"],
    "shop_number": ["shop_number", "do_kon_raqami", "do'kon_raqami"],
    "shop_owner_full_name": ["shop_owner_full_name", "do_kon_egasi_f_i_sh", "do'kon_egasi_f_i_sh"],
    "shop_owner_jshshir": ["shop_owner_jshshir", "jshshir_raqami"],
    "shop_owner_phone": ["shop_owner_phone", "telefon_raqami"],
    "shop_owner_company_name": ["shop_owner_company_name", "do_kon_egasiga_tegishli_mchj_nomi", "do'kon_egasiga_tegishli_mchj_nomi"],
    "shop_total_area": ["shop_total_area", "do_kon_umumiy_maydoni", "do'kon_umumiy_maydoni"],
    "tenant_count": ["tenant_count", "ijaraga_olgan_tadbirkor_soni"],
    "rented_area_sqm": ["rented_area_sqm", "ijaraga_berilgan_maydoni_kv_m", "ijaraga_berilgan_maydoni"],
    "business_name": ["business_name", "faoliyat_olib_borayotgan_tadbirkorlik_subyektining_nomi"],
    "business_leader_jshshir": ["business_leader_jshshir", "faoliyat_olib_borayotgan_tadbirkorlik_subyekti_rahbarining_jshshir_raqami"],
    "business_leader_full_name": ["business_leader_full_name", "faoliyat_olib_borayotgan_tadbirkorlik_subyekti_rahbarining_f_i_sh"],
    "business_leader_phone": ["business_leader_phone", "faoliyat_olib_borayotgan_tadbirkorlik_subyekti_rahbarining_telefon_raqami"],
    "business_inn": ["business_inn", "faoliyat_olib_borayotgan_tadbirkorlik_subyektining_guvoxnoma_raqami_stir_inn", "guvoxnoma_raqami_stir_inn"],
    "employee_count": ["employee_count", "faoliyat_olib_borayotgan_tadbirkorlik_subyektida_ishlaydigan_xodimlar_soni"],
    "employee_full_name": ["employee_full_name", "faoliyat_olib_borayotgan_tadbirkorlik_subyektida_ishlaydigan_xodimlar_f_i_sh_raqami"],
    "employee_jshshir": ["employee_jshshir", "faoliyat_olib_borayotgan_tadbirkorlik_subyektida_ishlaydigan_xodimlar_jshshir_raqami"],
    "employee_phone": ["employee_phone", "faoliyat_olib_borayotgan_tadbirkorlik_subyektida_ishlaydigan_xodimlar_telefon_raqami"],
    "business_type": ["business_type", "tadbirkorlik_turi"],
    "tax_type": ["tax_type", "soliq_turi_qqs_aylanmadan_olinadigan_soliq"],
    "cash_register_number": ["cash_register_number", "kassa_apparat_raqami"],
    "yearly_revenue_cash_register": ["yearly_revenue_cash_register", "yil_boshidan_savdo_tushumi_onlayn_nazorat_kassa_mashinasi"],
    "yearly_revenue_invoice": ["yearly_revenue_invoice", "yil_boshidan_savdo_tushumi_elektron_hisob_faktura"],
    "yearly_revenue_qr": ["yearly_revenue_qr", "yil_boshidan_savdo_tushumi_qr_kod_orqali"],
    "monthly_revenue_cash_register": ["monthly_revenue_cash_register", "oy_boshidan_savdo_tushumi_onlayn_nazorat_kassa_mashinasi"],
    "monthly_revenue_invoice": ["monthly_revenue_invoice", "oy_boshidan_savdo_tushumi_elektron_hisob_faktura"],
    "monthly_revenue_qr": ["monthly_revenue_qr", "oy_boshidan_savdo_tushumi_qr_kod_orqali"],
    "daily_revenue_cash_register": ["daily_revenue_cash_register", "kun_boshidan_savdo_tushumi_onlayn_nazorat_kassa_mashinasi"],
    "daily_revenue_invoice": ["daily_revenue_invoice", "kun_boshidan_savdo_tushumi_elektron_hisob_faktura"],
    "daily_revenue_qr": ["daily_revenue_qr", "kun_boshidan_savdo_tushumi_qr_kod_orqali"],
    "monthly_receipt_count": ["monthly_receipt_count", "berilgan_oylik_cheklar_soni"],
    "daily_receipt_count": ["daily_receipt_count", "berilgan_kunlik_cheklar_soni"],
    "monthly_visitors": ["monthly_visitors", "savdo_do_koniga_1_oy_davomida_kirayotgan_fuqarolar_soni_videokuzatuv_orqali"],
    "daily_visitors": ["daily_visitors", "savdo_do_koniga_1_kun_davomida_kirayotgan_fuqarolar_soni_videokuzatuv_orqali"],
    "business_status": ["business_status", "tadbirkorlik_subyekti_holati_faol_nofaol"],
    "fire_safety_level": ["fire_safety_level", "yong_in_xavfsizlik_darajasi"],
    "fire_alarm": ["fire_alarm", "yong_indan_xabar_beruvchi_qurilma"],
    "fire_extinguisher": ["fire_extinguisher", "birlamchi_yong_in_o_chirish_vositasi_ognetushitel_turi_soni"],
    "red_category_reason": ["red_category_reason", "qizil_toifaga_kirgan_do_kon_qizilga_kirgan_sababi"],
}


def resolve_header_map(row_values):
    raw_keys = [normalize_key(v) for v in row_values]
    result = {}

    for canonical_key, aliases in HEADER_ALIASES.items():
        for idx, key in enumerate(raw_keys):
            if key in aliases:
                result[canonical_key] = idx
                break
    return result


def find_header_row(ws):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [clean_text(v) for v in row]
        normalized = [normalize_key(v) for v in values]

        if "block" in normalized and "shop_number" in normalized:
            return row_idx, resolve_header_map(values)

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [clean_text(v) for v in row]
        header_map = resolve_header_map(values)
        if "block" in header_map and "shop_number" in header_map:
            return row_idx, header_map

    return None, {}


def get_cell(row, header_map, key):
    idx = header_map.get(key)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def build_tenant_lookup(shop, name, leader_jshshir):
    if leader_jshshir:
        obj = ShopTenant.objects.filter(shop=shop, leader_jshshir=leader_jshshir).first()
        if obj:
            return obj

    if name:
        obj = ShopTenant.objects.filter(shop=shop, name=name).first()
        if obj:
            return obj

    return None


def build_error_payload(excel_row_index, exc):
    if isinstance(exc, RowImportError):
        return {
            "row": excel_row_index,
            "excel_column": exc.field_key,
            "excel_column_label": FIELD_LABELS.get(exc.field_key, exc.field_key),
            "model_field": exc.model_field,
            "error": exc.message,
            "value": exc.value,
        }

    return {
        "row": excel_row_index,
        "excel_column": None,
        "excel_column_label": None,
        "model_field": None,
        "error": str(exc),
        "value": None,
    }


# =========================================================
# MAIN TASK
# =========================================================

@shared_task(bind=True)
def import_shop_excel_task(self, file_path):
    created_shops = 0
    updated_shops = 0
    created_tenants = 0
    updated_tenants = 0
    created_employees = 0
    skipped_rows = 0
    errors = []

    if not os.path.exists(file_path):
        return {
            "status": "error",
            "message": "Fayl topilmadi.",
            "created_shops": 0,
            "updated_shops": 0,
            "created_tenants": 0,
            "updated_tenants": 0,
            "created_employees": 0,
            "skipped_rows": 0,
            "errors": [],
        }

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row, header_map = find_header_row(ws)
    if not header_row:
        return {
            "status": "error",
            "message": "Excel header topilmadi.",
            "created_shops": 0,
            "updated_shops": 0,
            "created_tenants": 0,
            "updated_tenants": 0,
            "created_employees": 0,
            "skipped_rows": 0,
            "errors": [],
        }

    start_row = header_row + 1

    for excel_row_index, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        try:
            block_type = normalize_block_type(get_cell(row, header_map, "block"))
            shop_number = normalize_shop_number(get_cell(row, header_map, "shop_number"))

            if not block_type or not shop_number:
                skipped_rows += 1
                continue

            shop_number = validate_max_length("shop_number", "Shop.shop_number", shop_number)

            owner_fio = validate_max_length(
                "shop_owner_full_name",
                "Shop.owner_fio",
                empty_or_zero_to_none(get_cell(row, header_map, "shop_owner_full_name"))
            )
            owner_jshshir = validate_max_length(
                "shop_owner_jshshir",
                "Shop.owner_jshshir",
                normalize_jshshir(get_cell(row, header_map, "shop_owner_jshshir"))
            )
            owner_phone = validate_max_length(
                "shop_owner_phone",
                "Shop.owner_phone",
                normalize_phone(get_cell(row, header_map, "shop_owner_phone"))
            )
            owner_company_name = validate_max_length(
                "shop_owner_company_name",
                "Shop.owner_company_name",
                empty_or_zero_to_none(get_cell(row, header_map, "shop_owner_company_name"))
            )

            shop_defaults = {
                "owner_fio": owner_fio,
                "owner_jshshir": owner_jshshir,
                "owner_phone": owner_phone,
                "owner_company_name": owner_company_name,
                "total_area": parse_decimal(get_cell(row, header_map, "shop_total_area")),
                "tenants_count": parse_int(get_cell(row, header_map, "tenant_count")) or 0,
            }

            with transaction.atomic():
                shop, shop_created = Shop.objects.update_or_create(
                    block_type=block_type,
                    shop_number=shop_number,
                    defaults=shop_defaults
                )

                if shop_created:
                    created_shops += 1
                else:
                    updated_shops += 1

                business_name = validate_max_length(
                    "business_name",
                    "ShopTenant.name",
                    empty_or_zero_to_none(get_cell(row, header_map, "business_name"))
                )
                leader_jshshir = validate_max_length(
                    "business_leader_jshshir",
                    "ShopTenant.leader_jshshir",
                    normalize_jshshir(get_cell(row, header_map, "business_leader_jshshir"))
                )
                leader_fio = validate_max_length(
                    "business_leader_full_name",
                    "ShopTenant.leader_fio",
                    empty_or_zero_to_none(get_cell(row, header_map, "business_leader_full_name"))
                )
                leader_phone = validate_max_length(
                    "business_leader_phone",
                    "ShopTenant.leader_phone",
                    normalize_phone(get_cell(row, header_map, "business_leader_phone"))
                )

                if not business_name and not leader_fio and not leader_jshshir:
                    continue

                raw_business_inn = get_cell(row, header_map, "business_inn")
                stir, certificate_number = split_certificate_and_stir(raw_business_inn)

                stir = validate_max_length("business_inn", "ShopTenant.stir", stir)
                certificate_number = validate_max_length("business_inn", "ShopTenant.certificate_number", certificate_number)

                raw_cash_register = validate_max_length(
                    "cash_register_number",
                    "ShopTenant.cash_register_number_vat",
                    empty_or_zero_to_none(get_cell(row, header_map, "cash_register_number"))
                )

                extinguisher_info = validate_max_length(
                    "fire_extinguisher",
                    "ShopTenant.extinguisher_info",
                    empty_or_zero_to_none(get_cell(row, header_map, "fire_extinguisher"))
                )

                red_reason = validate_max_length(
                    "red_category_reason",
                    "ShopTenant.red_reason",
                    empty_or_zero_to_none(get_cell(row, header_map, "red_category_reason"))
                )

                tax_type_val = normalize_tax_type(get_cell(row, header_map, "tax_type"))
                is_vat = tax_type_val == ShopTenant.TaxType.VAT

                ytd_okkm = parse_decimal(get_cell(row, header_map, "yearly_revenue_cash_register"))
                ytd_e_payment = parse_decimal(get_cell(row, header_map, "yearly_revenue_invoice"))
                mtd_okkm = parse_decimal(get_cell(row, header_map, "monthly_revenue_cash_register"))
                mtd_e_payment = parse_decimal(get_cell(row, header_map, "monthly_revenue_invoice"))
                dtd_okkm = parse_decimal(get_cell(row, header_map, "daily_revenue_cash_register"))
                dtd_e_payment = parse_decimal(get_cell(row, header_map, "daily_revenue_invoice"))
                monthly_checks = parse_int(get_cell(row, header_map, "monthly_receipt_count")) or 0
                daily_checks = parse_int(get_cell(row, header_map, "daily_receipt_count")) or 0

                tenant_defaults = {
                    "rented_area": parse_decimal(get_cell(row, header_map, "rented_area_sqm")),
                    "business_type": normalize_business_type(get_cell(row, header_map, "business_type")),
                    "tax_type": tax_type_val,
                    "activity_status": normalize_activity_status(get_cell(row, header_map, "business_status")),
                    "name": business_name,
                    "leader_fio": leader_fio,
                    "leader_jshshir": leader_jshshir,
                    "leader_phone": leader_phone,
                    "stir": stir,
                    "certificate_number": certificate_number,
                    "employees_count": parse_int(get_cell(row, header_map, "employee_count")) or 0,
                    "cash_register_number_vat": raw_cash_register if is_vat else None,
                    "cash_register_number_turnover": None if is_vat else raw_cash_register,
                    "ytd_okkm_vat": ytd_okkm if is_vat else None,
                    "ytd_okkm_turnover": None if is_vat else ytd_okkm,
                    "ytd_e_payment_vat": ytd_e_payment if is_vat else None,
                    "ytd_e_payment_turnover": None if is_vat else ytd_e_payment,
                    "mtd_okkm_vat": mtd_okkm if is_vat else None,
                    "mtd_okkm_turnover": None if is_vat else mtd_okkm,
                    "mtd_e_payment_vat": mtd_e_payment if is_vat else None,
                    "mtd_e_payment_turnover": None if is_vat else mtd_e_payment,
                    "dtd_okkm_vat": dtd_okkm if is_vat else None,
                    "dtd_okkm_turnover": None if is_vat else dtd_okkm,
                    "dtd_e_payment_vat": dtd_e_payment if is_vat else None,
                    "dtd_e_payment_turnover": None if is_vat else dtd_e_payment,
                    "monthly_checks_count_vat": monthly_checks if is_vat else None,
                    "monthly_checks_count_turnover": None if is_vat else monthly_checks,
                    "daily_checks_count_vat": daily_checks if is_vat else None,
                    "daily_checks_count_turnover": None if is_vat else daily_checks,
                    "monthly_visitors": parse_int(get_cell(row, header_map, "monthly_visitors")) or 0,
                    "daily_visitors": parse_int(get_cell(row, header_map, "daily_visitors")) or 0,
                    "fire_safety_level": normalize_fire_safety_level(get_cell(row, header_map, "fire_safety_level")),
                    "has_fire_alarm": parse_bool(get_cell(row, header_map, "fire_alarm")),
                    "extinguisher_info": extinguisher_info,
                    "is_red_category": bool(red_reason),
                    "red_reason": red_reason,
                }

                tenant = build_tenant_lookup(shop=shop, name=business_name, leader_jshshir=leader_jshshir)

                if tenant:
                    for field, val in tenant_defaults.items():
                        setattr(tenant, field, val)
                    tenant.shop = shop
                    tenant.save()
                    updated_tenants += 1
                else:
                    tenant = ShopTenant.objects.create(shop=shop, **tenant_defaults)
                    created_tenants += 1

                employee_names = split_multiline(get_cell(row, header_map, "employee_full_name"))
                employee_jshshirs = split_multiline(get_cell(row, header_map, "employee_jshshir"))
                employee_phones = split_employee_phones(get_cell(row, header_map, "employee_phone"))

                tenant.employees.all().delete()

                max_len = max(len(employee_names), len(employee_jshshirs), len(employee_phones), 0)

                for i in range(max_len):
                    fio = employee_names[i] if i < len(employee_names) else None
                    jshshir = employee_jshshirs[i] if i < len(employee_jshshirs) else None
                    phone = employee_phones[i] if i < len(employee_phones) else None

                    fio = validate_max_length("employee_full_name", "TenantEmployee.fio", empty_or_zero_to_none(fio))
                    jshshir = validate_max_length("employee_jshshir", "TenantEmployee.jshshir", normalize_jshshir(jshshir))
                    phone = validate_max_length("employee_phone", "TenantEmployee.phone", normalize_phone(phone))

                    # 0 yoki pustoy bo'lsa create qilinmaydi
                    if not fio and not jshshir and not phone:
                        continue

                    TenantEmployee.objects.create(
                        tenant=tenant,
                        fio=fio,
                        jshshir=jshshir,
                        phone=phone,
                    )
                    created_employees += 1

        except Exception as exc:
            skipped_rows += 1
            errors.append(build_error_payload(excel_row_index, exc))

    return {
        "status": "success",
        "message": "Import yakunlandi.",
        "created_shops": created_shops,
        "updated_shops": updated_shops,
        "created_tenants": created_tenants,
        "updated_tenants": updated_tenants,
        "created_employees": created_employees,
        "skipped_rows": skipped_rows,
        "errors": errors,
    }


@shared_task(name="celery_task.sync_tenants_rent_task")
def sync_tenants_rent_task():
    """Ijara (rent) integratsiyasidan ijarachilarni ShopTenant ga sync qiladi (kunlik)."""
    from monitoring.services.rent_sync import sync_tenants_from_rent
    stats = sync_tenants_from_rent()
    return {
        "ok": True,
        "shops": stats.shops,
        "with_owner_cadastr": stats.with_owner_cadastr,
        "created": stats.tenants_created,
        "updated": stats.tenants_updated,
        "errors": stats.errors,
    }


@shared_task(bind=True, name="celery_task.sync_factura_revenue_task")
def sync_factura_revenue_task(self, period_from=None, period_to=None):
    """
    Faktura tushumini FacturaRevenueDaily ga yig'adi (dashboard grafigi uchun).
    Davriy ishlatish uchun (celery beat).
    """
    stats = sync_factura_revenue(period_from=period_from, period_to=period_to)
    return {
        "ok": True,
        "tenants": stats.tenants,
        "with_tin": stats.with_tin,
        "days_written": stats.days_written,
        "errors": stats.errors,
    }


@shared_task(name="celery_task.sync_tenant_soliq_task")
def sync_tenant_soliq_task():
    """
    Soliq rekvizit va faktura tushumini ShopTenant maydonlariga yozadi (FON da).

    Shu task tufayli shop-tenant API lari (ro'yxat/detal) request thread da
    soliqqa bormaydi — bazadan o'qiydi, boshqa API lar "pending" qolmaydi.
    Source of truth = soliq.
    """
    from monitoring.services.tenant_soliq_sync import sync_tenant_soliq
    stats = sync_tenant_soliq()
    return {
        "ok": True,
        "tenants": stats.tenants,
        "updated": stats.updated,
        "errors": stats.errors,
    }


@shared_task(name="celery_task.warm_shop_cash_status_task")
def warm_shop_cash_status_task():
    """
    Do'kon kassa (NKM) rangi keshini oldindan to'ldiradi (jonli soliqqa boradi).

    /shop/cash-status/ endpoint i FAQAT keshdan o'qiydi (request thread ni
    bloklamaslik uchun) — shu task kesh hali bo'sh do'konlarni "pending" dan
    yashil/qizilga o'tkazadi. Celery beat orqali davriy ishlaydi.
    """
    from monitoring.services.shop_status import shop_cash_status

    shops = Shop.objects.prefetch_related("tenants").all()
    counts = {"green": 0, "yellow": 0, "red": 0, "none": 0, "pending": 0}
    for shop in shops:
        st = shop_cash_status(shop)  # cache_only=False -> jonli, keshni to'ldiradi
        counts[st["color"] or "none"] += 1
    return {"ok": True, **counts}


@shared_task
def sync_patrol_cars_task():
    api_url = "http://25.1.1.217:80/api/mobject/lastData"
    token = "6bq6kimsmhniuothjbkai715n2"
    timeout = 20

    if not api_url:
        return {"ok": False, "error": "PATROL_CAR_LASTDATA_URL is not set"}

    stats = sync_patrol_cars_from_api(api_url=api_url, timeout=timeout, token=token)
    return {
        "ok": True,
        "fetched": stats.fetched,
        "created": stats.created,
        "updated": stats.updated,
        "skipped_no_imei": stats.skipped_no_imei,
        "errors": stats.errors,
    }


def normalize_pinfl(value):
    value = (value or "").strip()
    if not value:
        return None

    digits = "".join(ch for ch in value if ch.isdigit())
    if not digits:
        return None

    # faqat 0 lardan iborat bo'lsa yaroqsiz
    if set(digits) == {"0"}:
        return None

    return digits


@shared_task(bind=True, name="celery_task.sync_employee_avatars_from_long_term")
def sync_employee_avatars_from_long_term(self) -> Dict[str, Any]:
    """
    Background task:
    TenantEmployee -> employee.jshshir -> external API -> image(file) -> employee.avatar
    """
    qs = TenantEmployee.objects.all()

    results: List[Dict[str, Any]] = []
    saved = 0
    failed = 0
    skipped = 0

    idx = 0
    for employee in qs:
        idx += 1
        pinfl = normalize_pinfl(employee.jshshir)

        if employee.avatar:
            skipped += 1
            print(
                f"[SYNC_AVATAR] ({idx}) SKIP allaqachon avatar bor employee_id={employee.id}",
                flush=True
            )
            results.append({
                "pinfl": pinfl,
                "employee_id": employee.id,
                "ok": True,
                "status": "skipped",
                "avatar": str(employee.avatar),
                "detail": "Avatar allaqachon mavjud",
            })
            continue

        if not pinfl:
            skipped += 1
            print(
                f"[SYNC_AVATAR] ({idx}) SKIP jshshir yo'q yoki noto'g'ri employee_id={employee.id}",
                flush=True
            )
            results.append({
                "pinfl": None,
                "employee_id": employee.id,
                "ok": False,
                "status": "skipped",
                "avatar": None,
                "detail": "JSHSHIR bo'sh yoki noto'g'ri",
            })
            continue

        try:
            fetched = fetch_citizen_avatar_file(pinfl)

            if not fetched:
                failed += 1
                results.append({
                    "pinfl": pinfl,
                    "employee_id": employee.id,
                    "ok": False,
                    "status": "failed",
                    "avatar": None,
                    "detail": "Rasm topilmadi yoki servis xato qaytardi",
                })
                continue

            filename, file_obj = fetched

            employee.avatar.save(filename, file_obj, save=True)

            saved += 1
            print(
                f"[SYNC_AVATAR] ({idx}) SAVE ok avatar={employee.avatar.name} employee_id={employee.id}",
                flush=True
            )
            results.append({
                "pinfl": pinfl,
                "employee_id": employee.id,
                "ok": True,
                "status": "saved",
                "avatar": employee.avatar.name,
                "detail": "Saqlandi",
            })

        except Exception as e:
            failed += 1
            print(
                f"[SYNC_AVATAR] ({idx}) ERROR pinfl={pinfl} employee_id={employee.id} err={type(e).__name__}: {e}",
                flush=True
            )
            results.append({
                "pinfl": pinfl,
                "employee_id": employee.id,
                "ok": False,
                "status": "failed",
                "avatar": None,
                "detail": f"Xatolik: {type(e).__name__}: {e}",
            })

    return {
        "ok": True,
        "saved": saved,
        "failed": failed,
        "skipped": skipped,
        "total": qs.count(),
        "results": results,
    }

@shared_task(bind=True, name="celery_task.sync_shop_owner_avatars_task")
def sync_shop_owner_avatars_task(self) -> Dict[str, Any]:
    """
    Shop.owner_jshshir -> external API -> Shop.avatar
    """
    qs = Shop.objects.all()

    results: List[Dict[str, Any]] = []
    saved = 0
    failed = 0
    skipped = 0

    idx = 0
    for shop in qs:
        idx += 1
        pinfl = normalize_pinfl(shop.owner_jshshir)

        if shop.avatar:
            skipped += 1
            print(
                f"[SYNC_SHOP_AVATAR] ({idx}) SKIP avatar bor shop_id={shop.id}",
                flush=True
            )
            results.append({
                "shop_id": shop.id,
                "shop_number": shop.shop_number,
                "block_type": shop.block_type,
                "pinfl": pinfl,
                "ok": True,
                "status": "skipped",
                "avatar": str(shop.avatar),
                "detail": "Avatar allaqachon mavjud",
            })
            continue

        if not pinfl:
            skipped += 1
            print(
                f"[SYNC_SHOP_AVATAR] ({idx}) SKIP owner_jshshir yo'q shop_id={shop.id}",
                flush=True
            )
            results.append({
                "shop_id": shop.id,
                "shop_number": shop.shop_number,
                "block_type": shop.block_type,
                "pinfl": None,
                "ok": False,
                "status": "skipped",
                "avatar": None,
                "detail": "owner_jshshir bo'sh yoki noto'g'ri",
            })
            continue

        try:
            fetched = fetch_citizen_avatar_file(pinfl)

            if not fetched:
                failed += 1
                results.append({
                    "shop_id": shop.id,
                    "shop_number": shop.shop_number,
                    "block_type": shop.block_type,
                    "pinfl": pinfl,
                    "ok": False,
                    "status": "failed",
                    "avatar": None,
                    "detail": "Rasm topilmadi yoki servis xato qaytardi",
                })
                continue

            filename, file_obj = fetched
            shop.avatar.save(filename, file_obj, save=True)

            saved += 1
            print(
                f"[SYNC_SHOP_AVATAR] ({idx}) SAVE ok avatar={shop.avatar.name} shop_id={shop.id}",
                flush=True
            )
            results.append({
                "shop_id": shop.id,
                "shop_number": shop.shop_number,
                "block_type": shop.block_type,
                "pinfl": pinfl,
                "ok": True,
                "status": "saved",
                "avatar": shop.avatar.name,
                "detail": "Shop owner avatar saqlandi",
            })

        except Exception as e:
            failed += 1
            print(
                f"[SYNC_SHOP_AVATAR] ({idx}) ERROR pinfl={pinfl} shop_id={shop.id} err={type(e).__name__}: {e}",
                flush=True
            )
            results.append({
                "shop_id": shop.id,
                "shop_number": shop.shop_number,
                "block_type": shop.block_type,
                "pinfl": pinfl,
                "ok": False,
                "status": "failed",
                "avatar": None,
                "detail": f"Xatolik: {type(e).__name__}: {e}",
            })

    return {
        "ok": True,
        "saved": saved,
        "failed": failed,
        "skipped": skipped,
        "total": qs.count(),
        "results": results,
    }


@shared_task(bind=True, name="celery_task.sync_shop_tenant_leader_avatars_task")
def sync_shop_tenant_leader_avatars_task(self) -> Dict[str, Any]:
    """
    ShopTenant.leader_jshshir -> external API -> ShopTenant.avatar
    """
    qs = ShopTenant.objects.select_related("shop").all()

    results: List[Dict[str, Any]] = []
    saved = 0
    failed = 0
    skipped = 0

    idx = 0
    for tenant in qs:
        idx += 1
        pinfl = normalize_pinfl(tenant.leader_jshshir)

        if tenant.avatar:
            skipped += 1
            print(
                f"[SYNC_TENANT_AVATAR] ({idx}) SKIP avatar bor tenant_id={tenant.id}",
                flush=True
            )
            results.append({
                "tenant_id": tenant.id,
                "tenant_name": tenant.name,
                "shop_id": tenant.shop_id,
                "shop_number": tenant.shop.shop_number if tenant.shop else None,
                "pinfl": pinfl,
                "ok": True,
                "status": "skipped",
                "avatar": str(tenant.avatar),
                "detail": "Avatar allaqachon mavjud",
            })
            continue

        if not pinfl:
            skipped += 1
            print(
                f"[SYNC_TENANT_AVATAR] ({idx}) SKIP leader_jshshir yo'q tenant_id={tenant.id}",
                flush=True
            )
            results.append({
                "tenant_id": tenant.id,
                "tenant_name": tenant.name,
                "shop_id": tenant.shop_id,
                "shop_number": tenant.shop.shop_number if tenant.shop else None,
                "pinfl": None,
                "ok": False,
                "status": "skipped",
                "avatar": None,
                "detail": "leader_jshshir bo'sh yoki noto'g'ri",
            })
            continue

        try:
            fetched = fetch_citizen_avatar_file(pinfl)

            if not fetched:
                failed += 1
                results.append({
                    "tenant_id": tenant.id,
                    "tenant_name": tenant.name,
                    "shop_id": tenant.shop_id,
                    "shop_number": tenant.shop.shop_number if tenant.shop else None,
                    "pinfl": pinfl,
                    "ok": False,
                    "status": "failed",
                    "avatar": None,
                    "detail": "Rasm topilmadi yoki servis xato qaytardi",
                })
                continue

            filename, file_obj = fetched
            tenant.avatar.save(filename, file_obj, save=True)

            saved += 1
            print(
                f"[SYNC_TENANT_AVATAR] ({idx}) SAVE ok avatar={tenant.avatar.name} tenant_id={tenant.id}",
                flush=True
            )
            results.append({
                "tenant_id": tenant.id,
                "tenant_name": tenant.name,
                "shop_id": tenant.shop_id,
                "shop_number": tenant.shop.shop_number if tenant.shop else None,
                "pinfl": pinfl,
                "ok": True,
                "status": "saved",
                "avatar": tenant.avatar.name,
                "detail": "ShopTenant leader avatar saqlandi",
            })

        except Exception as e:
            failed += 1
            print(
                f"[SYNC_TENANT_AVATAR] ({idx}) ERROR pinfl={pinfl} tenant_id={tenant.id} err={type(e).__name__}: {e}",
                flush=True
            )
            results.append({
                "tenant_id": tenant.id,
                "tenant_name": tenant.name,
                "shop_id": tenant.shop_id,
                "shop_number": tenant.shop.shop_number if tenant.shop else None,
                "pinfl": pinfl,
                "ok": False,
                "status": "failed",
                "avatar": None,
                "detail": f"Xatolik: {type(e).__name__}: {e}",
            })

    return {
        "ok": True,
        "saved": saved,
        "failed": failed,
        "skipped": skipped,
        "total": qs.count(),
        "results": results,
    }