# cameras/views.py
import re
from io import BytesIO

from django.db import transaction
from django.utils.translation import gettext_lazy as _
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status

from openpyxl import load_workbook

from directory.models import Mahalla  # sizda shu path bo‘ladi
from monitoring.models import CameraInformation
from monitoring.serializers import CameraInformationImportSerializer


def _norm_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _parse_location(value: str):
    """
    Excel 'location' ustuni: "LAT, LON"
    Returns: (lat_str, lon_str) or (None, None)
    """
    s = _norm_str(value)
    if not s:
        return None, None

    # "41.32, 69.24" yoki "41.32 69.24" kabi holatlar
    s = s.replace(";", ",")
    s = re.sub(r"\s+", " ", s).strip()

    if "," in s:
        parts = [p.strip() for p in s.split(",") if p.strip()]
    else:
        parts = [p.strip() for p in s.split(" ") if p.strip()]

    if len(parts) < 2:
        return None, None

    lat, lon = parts[0], parts[1]

    # oddiy float tekshiruv (string ko‘rinishda saqlayapmiz)
    try:
        float(lat)
        float(lon)
    except Exception:
        return None, None

    return lat, lon


class CameraInformationImportInfoView(APIView):
    """
    POST multipart/form-data:
      - mahalla_id: int
      - file: .xlsx

    URL:
      path('camera-information/fields/import-by-mahalla-id/', CameraInformationImportInfoView.as_view(), ...)
    """
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        ser = CameraInformationImportSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        mahalla_id = ser.validated_data["mahalla_id"]
        excel_file = ser.validated_data["file"]

        try:
            mahalla = Mahalla.objects.select_related("district", "district__region").get(id=mahalla_id)
        except Mahalla.DoesNotExist:
            return Response(
                {"detail": f"Mahalla topilmadi: {mahalla_id}"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Excel o‘qish
        try:
            wb = load_workbook(filename=BytesIO(excel_file.read()), data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"detail": f"Excel o‘qishda xatolik: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Header topish: 1-qatorni header deb olamiz (sizning shablonda shunday)
        # Kutilgan ustunlar:
        expected = {
            "no": ["no", "№", "#"],
            "object_name": ["object_name", "obyekt", "object", "name"],
            "address": ["address", "Address"],
            "ip_address": ["ip_address", "ip", "ip address"],
            "login": ["login"],
            "parol": ["parol", "password", "пароль"],
            "camera_type": ["camera_type", "type"],
            "direction": ["direction", "yo'nalish", "йўналиш"],
            "location": ["location", "koordinata", "coord", "координата"],
        }

        def normalize_header(h):
            h = _norm_str(h).lower()
            h = h.replace("\n", " ").strip()
            return h

        header_row = None
        header_map = {}  # field -> col_index(1-based)

        # 1..5 qator orasidan header qidiramiz (ba’zi excelda tepada title bo‘lishi mumkin)
        for r in range(1, 6):
            row_vals = [normalize_header(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
            if not any(row_vals):
                continue

            # shu qatorda kutilganlardan kamida 3 tasi topilsa header deb olamiz
            tmp_map = {}
            for c, hv in enumerate(row_vals, start=1):
                if not hv:
                    continue
                for key, variants in expected.items():
                    if key in tmp_map:
                        continue
                    if any(v in hv for v in variants):
                        tmp_map[key] = c

            if len(tmp_map) >= 3:
                header_row = r
                header_map = tmp_map
                break

        if not header_row:
            return Response(
                {"detail": "Header topilmadi. Ustun nomlari shablondagi kabi bo‘lishi kerak."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Majburiy ustunlar
        required_cols = ["object_name", "ip_address", "location"]
        missing = [k for k in required_cols if k not in header_map]
        if missing:
            return Response(
                {"detail": f"Majburiy ustunlar topilmadi: {missing}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        start_row = header_row + 1

        created = 0
        updated = 0
        skipped = 0
        errors = []

        # region/district avtomatik (agar modelda mavjud bo‘lsa)
        mahalla_district = getattr(mahalla, "district", None)
        mahalla_region = getattr(mahalla_district, "region", None) if mahalla_district else None

        # Import: hammasini bitta transactionda
        with transaction.atomic():
            for r in range(start_row, ws.max_row + 1):
                def cell(key):
                    col = header_map.get(key)
                    if not col:
                        return None
                    return ws.cell(row=r, column=col).value

                object_name = _norm_str(cell("object_name"))
                address = _norm_str(cell("address"))
                ip_address = _norm_str(cell("ip_address"))
                login = _norm_str(cell("login"))
                parol = _norm_str(cell("parol"))
                camera_type = _norm_str(cell("camera_type"))
                direction = _norm_str(cell("direction"))
                location = cell("location")

                # bo‘sh qatorlarni tashlab ketish
                if not any([object_name, ip_address, address, _norm_str(location), direction, login, parol, camera_type]):
                    skipped += 1
                    continue

                if not ip_address:
                    errors.append({"row": r, "error": "ip_address bo‘sh", "data": {"object_name": object_name}})
                    continue

                lat, lon = _parse_location(location)
                if not lat or not lon:
                    errors.append({"row": r, "error": "location noto‘g‘ri. Format: 'LAT, LON'", "data": {"ip_address": ip_address, "location": _norm_str(location)}})
                    continue

                defaults = {
                    "object_name": object_name or None,
                    "address": address or None,
                    "direction": direction or None,
                    "ip_address": ip_address,
                    "mahalla": mahalla,
                    "coordinate_x": lat,  # sizda coordinate_x CharField
                    "coordinate_y": lon,
                    "login": login or None,
                    "parol": parol or None,
                    "camera_type": camera_type or None,
                }

                # status default
                defaults["status"] = CameraInformation.STATUS.INSTALLED

                # district/region set (agar mavjud bo‘lsa)
                if hasattr(CameraInformation, "district_id") and mahalla_district:
                    defaults["district"] = mahalla_district
                if hasattr(CameraInformation, "region_id") and mahalla_region:
                    defaults["region"] = mahalla_region

                obj, is_created = CameraInformation.objects.update_or_create(
                    ip_address=ip_address,
                    defaults=defaults,
                )

                if is_created:
                    created += 1
                else:
                    updated += 1

        return Response(
            {
                "detail": "Import yakunlandi",
                "mahalla_id": mahalla_id,
                "header_row": header_row,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors_count": len(errors),
                "errors": errors[:200],  # juda ko‘p bo‘lsa response og‘irlashmasin
            },
            status=status.HTTP_200_OK,
        )
